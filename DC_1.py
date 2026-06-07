Created on Sun Mar 17 04:19:53 2025

@author: KRISYA YA
"""

# IMPORTS AND BASIC INFORMATION

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import openpyxl
import re


excel_file = 'elka_30_08_252.xlsx' # REPLACE WITH YOUR FILE NAME
first_time = float(7205.17) # REPLACE WITH TOTAL TIME FROM THE FIRST PROTOCOL
end_time = float(14400 - first_time)
protocol_number = 2 #CHANGE TO PROTOCOL NUMBER 1 OR 2
kit_number = 4 # REPLACE WITH THE NUMBER OF KITTENS IN THE LITTER 1-5

workbook = load_workbook(excel_file)
sheet = workbook.active

print(f"Активный лист: {sheet.title}")

# MAIN TABLE AND UPPER TABLES
words_string = "MKit, KitM, KitKit, TotalKit, Total, M, 1, 2, 3, 4, 5, 1init, 1rec, 2init, 2rec, 3init, 3rec, 4init, 4rec, 5init, 5rec, 12, 13, 14, 15, 21, 23, 24, 25, 31, 32, 34, 35, 41, 42, 43, 45, 51, 52, 53, 54, 1M, 2M, 3M, 4M, 5M, M1, M2, M3, M4, M5, 1neighbour, 2neighbour, 3neighbour, 4neighbour, 5neighbour, KitNeighbour, Neighbour1, Neighbour2, Neighbour3, Neighbour4, Neighbour5, NeighbourKit, Mneighbour, NeighbourM" 


column_letter = 'M'
sheet.column_dimensions[column_letter].width = 15

words = [word.strip() for word in words_string.split(',')]
for i, word in enumerate(words):
    row = i + 6
    cell = f"M{row}"
    sheet[cell] = word

workbook.save(excel_file)



words_string = "cocanie, nevidno, allogruming, gruming, igra, selfplay, igrasmamojAct, igrasmamojPas, spredmetom, bokom, ckradivanie, zataivanie, lapki, nabeg, naprig, obxvat, pogonya, prigl" 

words = [word.strip() for word in words_string.split(',')]
start_column = 14

for i, word in enumerate(words):
    column = start_column + i
    cell = sheet.cell(row=5, column=column)
    cell.value = word
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    column_letter = get_column_letter(column)
    sheet.column_dimensions[column_letter].width = len(word) + 4  # +2 для небольшого отст

workbook.save(excel_file)



start_cell = "N6"
end_cell = "AE70"

thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])
    raise ValueError(f"Невозможно извлечь номер строки из адреса: {cell_address}")

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))  
start_row = extract_row_number(start_cell)  
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))  
end_row = extract_row_number(end_cell)  

for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border

workbook.save(excel_file)



start_cell = "M5"
end_cell = "AE5"

medium_border = Border(
    left=Side(border_style="medium", color="000000"),
    right=Side(border_style="medium", color="000000"),
    top=Side(border_style="medium", color="000000"),
    bottom=Side(border_style="medium", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])
    raise ValueError(f"Невозможно извлечь номер строки из адреса: {cell_address}")

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
start_row = extract_row_number(start_cell)
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
end_row = extract_row_number(end_cell)

for col in range(start_col, end_col + 1):
    cell = sheet.cell(row=start_row, column=col)
    cell.border = medium_border

workbook.save(excel_file)



start_cell = "M5"
end_cell = "M70"

medium_border = Border(
    left=Side(border_style="medium", color="000000"),
    right=Side(border_style="medium", color="000000"),
    top=Side(border_style="medium", color="000000"),
    bottom=Side(border_style="medium", color="000000")
)

start_col = column_index_from_string(start_cell[0])
start_row = int(start_cell[1:])
end_col = column_index_from_string(end_cell[0])
end_row = int(end_cell[1:])

for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = medium_border

workbook.save(excel_file)



fill_color = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")

start_row = 6  
end_row = 70
start_col = column_index_from_string("M")  
end_col = column_index_from_string("AE")   

for row in range(start_row, end_row + 1):
    if row % 2 == 0:
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = fill_color

workbook.save(excel_file)

print(f"Таблица категорий построена.")


words_string = "Mstop, Kitstop, success, unsuccess, TotalTime, nevidno, observed" 

words = [word.strip() for word in words_string.split(',')]
start_column = 13

column_index = start_column

for i, word in enumerate(words):
    if column_index == 17:
        column_index += 1

    cell = sheet.cell(row=1, column=column_index)
    cell.value = word

    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    column_letter = get_column_letter(column_index)
    sheet.column_dimensions[column_letter].width = len(word) + 4  # +4 для небольшого отступа

    column_index += 1

workbook.save(excel_file)



start_cell = "M1"
end_cell = "P3"

thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])
    raise ValueError(f"Невозможно извлечь номер строки из адреса: {cell_address}")

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))  
start_row = extract_row_number(start_cell)  
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))  
end_row = extract_row_number(end_cell)  

for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border

workbook.save(excel_file)



start_cell = "R1"
end_cell = "T2"

thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])
    raise ValueError(f"Невозможно извлечь номер строки из адреса: {cell_address}")

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))  
start_row = extract_row_number(start_cell)  
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))  
end_row = extract_row_number(end_cell)  

for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border

workbook.save(excel_file)



start_cell = "M1"
end_cell = "P1"

medium_border = Border(
    left=Side(border_style="medium", color="000000"),
    right=Side(border_style="medium", color="000000"),
    top=Side(border_style="medium", color="000000"),
    bottom=Side(border_style="medium", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])
    raise ValueError(f"Невозможно извлечь номер строки из адреса: {cell_address}")

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
start_row = extract_row_number(start_cell)
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
end_row = extract_row_number(end_cell)

for col in range(start_col, end_col + 1):
    cell = sheet.cell(row=start_row, column=col)
    cell.border = medium_border

workbook.save(excel_file)



start_cell = "R1"
end_cell = "T1"

medium_border = Border(
    left=Side(border_style="medium", color="000000"),
    right=Side(border_style="medium", color="000000"),
    top=Side(border_style="medium", color="000000"),
    bottom=Side(border_style="medium", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])
    raise ValueError(f"Невозможно извлечь номер строки из адреса: {cell_address}")

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
start_row = extract_row_number(start_cell)
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
end_row = extract_row_number(end_cell)

for col in range(start_col, end_col + 1):
    cell = sheet.cell(row=start_row, column=col)
    cell.border = medium_border

workbook.save(excel_file)

print(f"Таблица сосания и невидно построена.")


words_string = "Total, KitKit, MKit, KitM" 


column_letter = 'AI'
sheet.column_dimensions[column_letter].width = 15

words = [word.strip() for word in words_string.split(',')]
for i, word in enumerate(words):
    row = i + 6
    cell = f"AI{row}"
    sheet[cell] = word

workbook.save(excel_file)

words_string = "lable pairs, time, groups, time" 

workbook = load_workbook(excel_file)
sheet = workbook.active

words = [word.strip() for word in words_string.split(',')]
start_column = 33

for i, word in enumerate(words):
    column = start_column + i
    cell = sheet.cell(row=5, column=column)
    cell.value = word
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    column_letter = get_column_letter(column)
    sheet.column_dimensions[column_letter].width = len(word) + 4

workbook.save(excel_file)

start_cell = "AG6"
end_cell = "AH70"

thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))  
start_row = extract_row_number(start_cell)  
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))  
end_row = extract_row_number(end_cell)  

for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border

workbook.save(excel_file)

start_cell = "AI6"
end_cell = "AJ9"

thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))  
start_row = extract_row_number(start_cell)  
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))  
end_row = extract_row_number(end_cell)  

for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border

workbook.save(excel_file)

start_cell = "AG5"
end_cell = "AJ5"

medium_border = Border(
    left=Side(border_style="medium", color="000000"),
    right=Side(border_style="medium", color="000000"),
    top=Side(border_style="medium", color="000000"),
    bottom=Side(border_style="medium", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
start_row = extract_row_number(start_cell)
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
end_row = extract_row_number(end_cell)

for col in range(start_col, end_col + 1):
    cell = sheet.cell(row=start_row, column=col)
    cell.border = medium_border

workbook.save(excel_file)

def extract_row_number(cell_address):
    match = re.search(r'\d+$', cell_address)
    if match:
        return int(match.group())

def extract_column_letter(cell_address):
    match = re.search(r'^[A-Za-z]+', cell_address)
    if match:
        return match.group()

start_cell = 'AI6'
end_cell = 'AI9'

try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    medium_border = Border(
        left=Side(border_style="medium", color="000000"),
        right=Side(border_style="medium", color="000000"),
        top=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000")
    )

    start_col = column_index_from_string(extract_column_letter(start_cell))
    start_row = extract_row_number(start_cell)
    end_col = column_index_from_string(extract_column_letter(end_cell))
    end_row = extract_row_number(end_cell)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = medium_border

    workbook.save(excel_file)
except Exception as e:
    print(f"Произошла ошибка: {e}")


from openpyxl.styles import PatternFill

fill_color = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")

start_row = 6  
end_row = 70
start_col = column_index_from_string("AG")  
end_col = column_index_from_string("AH")   

for row in range(start_row, end_row + 1):
    if row % 2 == 0:
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = fill_color

workbook.save(excel_file)


print(f"Таблица борьбы построена.")


words_string = "1, 2, 3, 4, ?" 

column_letter = 'AL'
sheet.column_dimensions[column_letter].width = 15

words = [word.strip() for word in words_string.split(',')]
for i, word in enumerate(words):
    row = i + 6
    cell = f"AL{row}"
    sheet[cell] = word

workbook.save(excel_file)

words_string = "Pair, Success, Unsuccess, Total" 

workbook = load_workbook(excel_file)
sheet = workbook.active

words = [word.strip() for word in words_string.split(',')]
start_column = 38

for i, word in enumerate(words):
    column = start_column + i
    cell = sheet.cell(row=5, column=column)
    cell.value = word
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    column_letter = get_column_letter(column)
    sheet.column_dimensions[column_letter].width = len(word) + 4

workbook.save(excel_file)


start_cell = "AM6"
end_cell = "AO10"

thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))  
start_row = extract_row_number(start_cell)  
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))  
end_row = extract_row_number(end_cell)  

for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border

workbook.save(excel_file)


start_cell = "AL5"
end_cell = "AO5"

medium_border = Border(
    left=Side(border_style="medium", color="000000"),
    right=Side(border_style="medium", color="000000"),
    top=Side(border_style="medium", color="000000"),
    bottom=Side(border_style="medium", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
start_row = extract_row_number(start_cell)
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
end_row = extract_row_number(end_cell)

for col in range(start_col, end_col + 1):
    cell = sheet.cell(row=start_row, column=col)
    cell.border = medium_border

workbook.save(excel_file)


def extract_row_number(cell_address):
    match = re.search(r'\d+$', cell_address)
    if match:
        return int(match.group())

def extract_column_letter(cell_address):
    match = re.search(r'^[A-Za-z]+', cell_address)
    if match:
        return match.group()

start_cell = 'AL6'
end_cell = 'AL10'

try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    medium_border = Border(
        left=Side(border_style="medium", color="000000"),
        right=Side(border_style="medium", color="000000"),
        top=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000")
    )

    start_col = column_index_from_string(extract_column_letter(start_cell))
    start_row = extract_row_number(start_cell)
    end_col = column_index_from_string(extract_column_letter(end_cell))
    end_row = extract_row_number(end_cell)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = medium_border

    workbook.save(excel_file)
except Exception as e:
    print(f"Произошла ошибка: {e}")
    

fill_color = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")

start_row = 6  
end_row = 10
start_col = column_index_from_string("AL")  
end_col = column_index_from_string("AO")   

for row in range(start_row, end_row + 1):
    if row % 2 == 0:
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = fill_color

workbook.save(excel_file)

print(f"Нижняя таблица сосания построена.")


words_string = "1, 2, 3, 4, 5, M, 1x2, 1x3, 1x4, 1x5, 2x3, 2x4, 2x5, 3x4, 3x5, 4x5" 

column_letter = 'AI'
sheet.column_dimensions[column_letter].width = 15

words = [word.strip() for word in words_string.split(',')]
for i, word in enumerate(words):
    row = i + 6
    cell = f"AQ{row}"
    sheet[cell] = word

workbook.save(excel_file)


words_string = "Animals, Nevidno, Observed" 

words = [word.strip() for word in words_string.split(',')]
start_column = 43

for i, word in enumerate(words):
    column = start_column + i
    cell = sheet.cell(row=5, column=column)
    cell.value = word
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    column_letter = get_column_letter(column)
    sheet.column_dimensions[column_letter].width = len(word) + 4

workbook.save(excel_file)


start_cell = "AR6"
end_cell = "AS21"

thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))  
start_row = extract_row_number(start_cell)  
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))  
end_row = extract_row_number(end_cell)  

for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border

workbook.save(excel_file)


start_cell = "AQ5"
end_cell = "AS5"

medium_border = Border(
    left=Side(border_style="medium", color="000000"),
    right=Side(border_style="medium", color="000000"),
    top=Side(border_style="medium", color="000000"),
    bottom=Side(border_style="medium", color="000000")
)

def extract_row_number(cell_address):
    for i, char in enumerate(cell_address):
        if char.isdigit():
            return int(cell_address[i:])

start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
start_row = extract_row_number(start_cell)
end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
end_row = extract_row_number(end_cell)

for col in range(start_col, end_col + 1):
    cell = sheet.cell(row=start_row, column=col)
    cell.border = medium_border

workbook.save(excel_file)


def extract_row_number(cell_address):
    match = re.search(r'\d+$', cell_address)
    if match:
        return int(match.group())

def extract_column_letter(cell_address):
    match = re.search(r'^[A-Za-z]+', cell_address)
    if match:
        return match.group()

start_cell = 'AQ6'
end_cell = 'AQ21'

try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    medium_border = Border(
        left=Side(border_style="medium", color="000000"),
        right=Side(border_style="medium", color="000000"),
        top=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000")
    )

    start_col = column_index_from_string(extract_column_letter(start_cell))
    start_row = extract_row_number(start_cell)
    end_col = column_index_from_string(extract_column_letter(end_cell))
    end_row = extract_row_number(end_cell)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = medium_border

    workbook.save(excel_file)
except Exception as e:
    print(f"Произошла ошибка: {e}")
    

fill_color = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")

start_row = 6  
end_row = 21
start_col = column_index_from_string("AQ")  
end_col = column_index_from_string("AS")   

for row in range(start_row, end_row + 1):
    if row % 2 == 0:
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = fill_color

workbook.save(excel_file)

print("Нижняя таблица невидно построена.")


#     1 PROTOCOL
if protocol_number == 1:
    print("ПРОТОКОЛ 1")
    
#     CATEGORIES
    action_column = 'D'
    animal_column = 'K'

    content_to_count = "allogrumin"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                        
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
               
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
               
            elif str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'P{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета аллогруминга записан в сотлбец P.")




    content_to_count = "gruming   "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                 
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'Q{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета груминга записан в сотлбец Q.")




    content_to_count = "igra      "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
                
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'R{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета игры записан в сотлбец R.")

     

    content_to_count = "igra      "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "11":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "22":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "33":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "44":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "55":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "MM":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'S{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета selfplay записан в сотлбец S.")




    content_to_count = "igrasmamoj"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
                           
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'T{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета игра с активной мамой записан в сотлбец T.")





    content_to_count = "igrasmamoj"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "1tail":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "1ears":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "1paw":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2ears":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "2tail":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "2paw":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3ears":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "3paw":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "3tail":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4ears":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "4tail":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "4paw":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5ears":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "5paw":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "5tail":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittentail":
                countKM += 1
            elif str(cell_additional_value.value) == "Kittentail":
                countKM += 1
            elif str(cell_additional_value.value) == "kittenears":
                countKM += 1
            elif str(cell_additional_value.value) == "Kittenears":
                countKM += 1
            elif str(cell_additional_value.value) == "kittenpaw":
                countKM += 1
            elif str(cell_additional_value.value) == "Kittenpaw":
                countKM += 1
                            
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'U{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета игра с пассивной мамой записан в сотлбец U.")

     


    content_to_count = "spredmetom"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                 
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'V{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета игра с предметом записан в сотлбец V.")

     


    content_to_count = "bokom     "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
                         
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
             
            elif str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                 
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'W{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета боком записан в сотлбец W.")




    content_to_count = "ckradivan "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
                         
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
             
            elif str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
          
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'X{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета обхват записан в сотлбец X.")

     


    content_to_count = "zataivanie"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
                         
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
              
            elif str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
          
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'Y{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета затаивание записан в сотлбец Y.")

     


    content_to_count = "lapki     "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M11M":
                countMK += 1
                countM1 += 1
                count1M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "M22M":
                countMK += 1
                countM2 += 1
                count2M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "M33M":
                countMK += 1
                countM3 += 1
                count3M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "M44M":
                countMK += 1
                countM4 += 1
                count4M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "M55M":
                countMK += 1
                countM5 += 1
                count5M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "MkittenkittemM":
                countMK += 1
                countKM += 1
            elif str(cell_additional_value.value) == "MKittenKittenM":
                countMK += 1
                countKM += 1
            elif str(cell_additional_value.value) == "MKittenkittenM":
                countMK += 1
                countKM += 1
            
            elif str(cell_additional_value.value) == "1MM1":
                countMK += 1
                countM1 += 1
                count1M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "2MM2":
                countMK += 1
                countM2 += 1
                count2M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "3MM3":
                countMK += 1
                countM3 += 1
                count3M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "4MM4":
                countMK += 1
                countM4 += 1
                count4M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "5MM5":
                countMK += 1
                countM5 += 1
                count5M += 1
                countKM += 1
            elif str(cell_additional_value.value) == "kittenMMkitten":
                countMK += 1
                countKM += 1
            elif str(cell_additional_value.value) == "KittenMMKitten":
                countMK += 1
                countKM += 1
            elif str(cell_additional_value.value) == "KittenMMkitten":
                countMK += 1
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbourneighbour1":
                count1N += 1
                countKN += 1
                countN1 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "2neighbourneighbour2":
                count2N += 1
                countKN += 1
                countN2 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "3neighbourneighbour3":
                count3N += 1
                countKN += 1
                countN3 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "4neighbourneighbour4":
                count4N += 1
                countKN += 1
                countN4 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "5neighbourneighbour5":
                count5N += 1
                countKN += 1
                countN5 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "MneighbourneighbourM":
                countMN += 1
                countNM += 1
            elif str(cell_additional_value.value) == "KittenneighbourneighbourKitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "kittenneighbourneighbourkitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "Kittenneighbourneighbourkitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "kittenneighbourneighbourKitten":
                countKN += 1
                countNK += 1
                
            elif str(cell_additional_value.value) == "neighbour11neighbour":
                count1N += 1
                countKN += 1
                countN1 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbour22neighbour":
                count2N += 1
                countKN += 1
                countN2 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbour33neighbour":
                count3N += 1
                countKN += 1
                countN3 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbour44neighbour":
                count4N += 1
                countKN += 1
                countN4 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbour55neighbour":
                count5N += 1
                countKN += 1
                countN5 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourMMneighbour":
                countNM += 1
                countMN += 1
            elif str(cell_additional_value.value) == "neighbourKittenKittenneighbour":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkittenkittenneighbour":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourKittenkittenneighbour":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkittenKittenneighbour":
                countKN += 1
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighborneighbor1":
                count1N += 1
                countKN += 1
                countN1 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "2neighborneighbor2":
                count2N += 1
                countKN += 1
                countN2 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "3neighborneighbor3":
                count3N += 1
                countKN += 1
                countN3 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "4neighborneighbor4":
                count4N += 1
                countKN += 1
                countN4 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "5neighborneighbor5":
                count5N += 1
                countKN += 1
                countN5 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "MneighborneighborM":
                countMN += 1
                countNM += 1
            elif str(cell_additional_value.value) == "KittenneighborneighborKitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "kittenneighborneighborkitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "Kittenneighborneighborkitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "kittenneighborneighborKitten":
                countKN += 1
                countNK += 1
                
            elif str(cell_additional_value.value) == "neighbor11neighbor":
                count1N += 1
                countKN += 1
                countN1 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbor22neighbor":
                count2N += 1
                countKN += 1
                countN2 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbor33neighbor":
                count3N += 1
                countKN += 1
                countN3 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbor44neighbor":
                count4N += 1
                countKN += 1
                countN4 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighbor55neighbor":
                count5N += 1
                countKN += 1
                countN5 += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighborMMneighbor":
                countNM += 1
                countMN += 1
            elif str(cell_additional_value.value) == "neighborKittenKittenneighbor":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkittenkittenneighbor":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighborKittenkittenneighbor":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkittenKittenneighbor":
                countKN += 1
                countNK += 1
                         
            elif str(cell_additional_value.value) == "1221":
                countKK += 1
                count12 += 1
                count21 += 1
                count1init += 1
                count1rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "1331":
                countKK += 1
                count13 += 1
                count31 += 1
                count1init += 1
                count1rec += 1
                count3init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "1441":
                countKK += 1
                count14 += 1
                count41 += 1
                count1init += 1
                count1rec += 1
                count4init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "1551":
                countKK += 1
                count15 += 1
                count51 += 1
                count1init += 1
                count1rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "2112":
                countKK += 1
                count12 += 1
                count21 += 1
                count1init += 1
                count1rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "2332":
                countKK += 1
                count32 += 1
                count23 += 1
                count3init += 1
                count3rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "2442":
                countKK += 1
                count42 += 1
                count24 += 1
                count4init += 1
                count4rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "2552":
                countKK += 1
                count52 += 1
                count25 += 1
                count5init += 1
                count5rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "3113":
                countKK += 1
                count31 += 1
                count13 += 1
                count1init += 1
                count1rec += 1
                count3init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "3223":
                countKK += 1
                count32 += 1
                count23 += 1
                count3init += 1
                count3rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "3443":
                countKK += 1
                count34 += 1
                count43 += 1
                count3init += 1
                count3rec += 1
                count4init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "3553":
                countKK += 1
                count35 += 1
                count53 += 1
                count3init += 1
                count3rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "4114":
                countKK += 1
                count41 += 1
                count14 += 1
                count1init += 1
                count1rec += 1
                count4init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "4224":
                countKK += 1
                count42 += 1
                count24 += 1
                count4init += 1
                count4rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "4334":
                countKK += 1
                count43 += 1
                count34 += 1
                count4init += 1
                count4rec += 1
                count3init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "4554":
                countKK += 1
                count45 += 1
                count54 += 1
                count4init += 1
                count4rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "5115":
                countKK += 1
                count51 += 1
                count15 += 1
                count1init += 1
                count1rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "5225":
                countKK += 1
                count52 += 1
                count25 += 1
                count5init += 1
                count5rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "5335":
                countKK += 1
                count53 += 1
                count35 += 1
                count3init += 1
                count3rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "5445":
                countKK += 1
                count54 += 1
                count45 += 1
                count5init += 1
                count5rec += 1
                count4init += 1
                count4rec += 1
                          
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'Z{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета лапки записан в сотлбец Z.")





    content_to_count = "nabeg     "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
                          
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
              
            elif str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AA{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета набег записан в сотлбец AA.")

     



    content_to_count = "naprigivan"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
                         
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AB{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета напрыгивание записан в сотлбец AB.")

     



    content_to_count = "obxvat    "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
                       
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
            
            elif str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AC{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета обхват записан в сотлбец AC.")

     


    content_to_count = "pogonya   "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
                          
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
              
            elif str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AD{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета погоня записан в сотлбец AD.")


    content_to_count = "priglashen"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value.value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value.value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value.value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value.value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value.value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value.value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value.value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value.value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value.value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value.value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value.value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value.value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value.value) == "KittenM":
                countKM += 1
            
            
            elif str(cell_additional_value.value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value.value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value.value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value.value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value.value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value.value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value.value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value.value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value.value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value.value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value.value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value.value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value.value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value.value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value.value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value.value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value.value) == "neighborkitten":
                countNK += 1
                       
            elif str(cell_additional_value.value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value.value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
             
            elif str(cell_additional_value.value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value.value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value.value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value.value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value.value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value.value) == "M":
                countM += 1
            elif str(cell_additional_value.value) == "kitten":
                countTK += 1
            elif str(cell_additional_value.value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value.value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value.value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value.value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value.value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value.value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value.value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value.value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value.value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value.value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value.value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value.value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value.value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value.value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value.value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value.value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AE{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета приглашение записан в сотлбец AE.")


    
    
#     КАТЕГОРИИ ТОТАЛ 1 ПРОТОКОЛ
    output_cell = "P10"
    column_letter = 'D'
    content_to_count = "allogrumin"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")

    content_to_count = "gruming   "
    output_cell = "Q10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")


    content_to_count = "igra      "
    output_cell = "R10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")

    content_to_count = "spredmetom"
    output_cell = "V10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")


    content_to_count = "bokom     "
    output_cell = "W10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")


    content_to_count = "ckradivan "
    output_cell = "X10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")


    content_to_count = "zataivanie"
    output_cell = "Y10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")


    content_to_count = "lapki     "
    output_cell = "Z10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")



    content_to_count = "nabeg     "
    output_cell = "AA10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")



    content_to_count = "naprigivan"
    output_cell = "AB10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")



    content_to_count = "obxvat    "
    output_cell = "AC10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")



    content_to_count = "pogonya   "
    output_cell = "AD10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")


    content_to_count = "priglashen"
    output_cell = "AE10"

    count = 0
    for row in sheet.iter_rows():
        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count#
    workbook.save(excel_file)#

    print(f"Общее количество '{content_to_count}': {count}, резултат записан в ячейку {output_cell}.")




    ws = workbook.active

    values = [ws['S11'].value, ws['S12'].value, ws['S13'].value, 
              ws['S14'].value, ws['S15'].value, ws['S16'].value]
    total = sum(v for v in values if v is not None)

    ws['S10'] = total
    workbook.save(excel_file)

    print(f"Общее количество selfplay: {total}, резултат записан в ячейку S10.")


    values = [ws['T6'].value, ws['T7'].value]
    total = sum(v for v in values if v is not None)

    ws['T10'] = total
    workbook.save(excel_file)

    print(f"Общее количество игра с активной мамой: {total}, резултат записан в ячейку T10.")



    values = [ws['U6'].value, ws['U7'].value]
    total = sum(v for v in values if v is not None)

    ws['U10'] = total
    workbook.save(excel_file)

    print(f"Общее количество игра с пассивной мамой: {total}, резултат записан в ячейку U10.")



    values = [ws['T10'].value, ws['U10'].value]
    total = sum(v for v in values if v is not None)

    ws['U4'] = total
    workbook.save(excel_file)


    variable = "igrasmamoj"

    ws['U3'] = variable
    workbook.save(excel_file)

    print(f"Общее количество игра c мамой (акт и пас): {total}, резултат записан в ячейку U4.")
    
#     NEVIDNO
#         LITTER OF 1 KITTEN
    if kit_number == 1:
        print("выводок из 1 котенка")
    
        output_cell = "S2"
    
        workbook = load_workbook(excel_file, data_only=True)
        sheet = workbook.active
    
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "1neighbor", "neighbour1", "neighbor1", "1neighbourneighbour1",
            "1neighborneighbor1", "neighbour11neighbour", "neighbor11neighbor",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }
        
        total_time = 0
        current_period_start = 0.0
        is_in_invisible_period = True
        last_time = None
        first_kit_event_found = False
        
        for row in sheet.iter_rows(min_row=7, values_only=True):
            time_val = row[5]  # колонка F
            if time_val is not None:
                try:
                    t = float(time_val)
                    if last_time is None or t > last_time:
                        last_time = t
                except (ValueError, TypeError):
                    continue
        
        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            action = row[3]
            animal = row[10] 
            time = row[5]
            
            if action is None or animal is None or time is None:
                continue
            
            action_str = str(action).strip().lower()
            animal_str = str(animal).strip().lower()
        
            try:
                current_time = float(time)
            except (ValueError, TypeError):
                continue
            
            is_kit_event = (animal_str == "1" or animal_str in kit1_aliases)
            
            if is_kit_event and not first_kit_event_found:
                first_kit_event_found = True
                
                if action_str != "nevidno" and action_str != "vokal":
                    total_time += current_time - current_period_start
                    is_in_invisible_period = False
                else:
                    pass
            
            if is_kit_event:
                if action_str == "nevidno":
                    if not is_in_invisible_period:
                        current_period_start = current_time
                        is_in_invisible_period = True
                else:
                    if is_in_invisible_period:
                        total_time += current_time - current_period_start
                        is_in_invisible_period = False
        
        if is_in_invisible_period and last_time is not None:
            total_time += last_time - current_period_start
        
        if not first_kit_event_found and last_time is not None:
            total_time = last_time
        
        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время невидно 1 котенка: {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")


#         LITTER OF 2 KITTEN
    elif kit_number == 2:
        print("выводок из 2 котят")
        
        output_cell = "S2"
        
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        workbook = load_workbook(excel_file)
        sheet = workbook.active
            
        total_time = 0
        kit1_nevidno = True
        kit2_nevidno = True
        simultaneous_nevidno_start_time = 0.0
        last_time = None
        
        for row in sheet.iter_rows(min_row=7, values_only=True):
            time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
            if time_val is not None:
                try:
                    t = float(time_val)
                    if last_time is None or t > last_time:
                        last_time = t
                except ValueError:
                    continue
        
        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
            "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
            "kitten1", "kitten1", "1kitten", "1kitten",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }
        
        kit2_aliases = {
            "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
            "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
            "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
            "kitten2", "kitten2", "2kitten", "2kitten",
            "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
            "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
            "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
            "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
            "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
        }
        
        both_kits_aliases = {
            "12", "21", "1221", "2112"
        }
        
        kit1_first_event_found = False
        kit2_first_event_found = False
        
        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
            animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if action is None or animal is None or time is None:
                continue

            action_str = str(action).strip().lower()
            animal_str = str(animal).strip().lower()
            
            try:
                current_time = float(time)
            except ValueError:
                continue
            
            if animal_str in kit1_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                else:
                    kit1_nevidno = False
                    
            elif animal_str in kit2_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                else:
                    kit2_nevidno = False
                    
            elif animal_str in both_kits_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit2_nevidno = True
                else:
                    kit1_nevidno = False
                    kit2_nevidno = False
            
            if kit1_nevidno and kit2_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    duration = current_time - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
    
        if simultaneous_nevidno_start_time is not None and last_time is not None:
            duration = last_time - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        
        
        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время одновременного невидно для 2 котят: {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")



#         LITTER OF 3 KITTEN
    elif kit_number == 3:
        print("выводок из 3 котят")
        
        output_cell = "S2"
        
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        workbook = load_workbook(excel_file)
        sheet = workbook.active
            
        total_time = 0
        kit1_nevidno = True
        kit2_nevidno = True
        kit3_nevidno = True
        simultaneous_nevidno_start_time = 0.0
        last_time = None
        
        for row in sheet.iter_rows(min_row=7, values_only=True):
            time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
            if time_val is not None:
                try:
                    t = float(time_val)
                    if last_time is None or t > last_time:
                        last_time = t
                except ValueError:
                    continue
        
        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
            "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
            "kitten1", "kitten1", "1kitten", "1kitten",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }
        
        kit2_aliases = {
            "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
            "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
            "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
            "kitten2", "kitten2", "2kitten", "2kitten",
            "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
            "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
            "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
            "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
            "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
        }
        
        kit3_aliases = {
            "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
            "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
            "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
            "kitten3", "kitten3", "3kitten", "3kitten",
            "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
            "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
            "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
            "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
            "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
        }
        
        kit12_aliases = {"12", "21", "1221", "2112"}
        kit13_aliases = {"13", "31", "1331", "3113"}
        kit23_aliases = {"23", "32", "2332", "3223"}
        
        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
            animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if action is None or animal is None or time is None:
                continue

            action_str = str(action).strip().lower()
            animal_str = str(animal).strip().lower()
            
            try:
                current_time = float(time)
            except ValueError:
                continue
            
            if animal_str in kit1_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                else:
                    kit1_nevidno = False
                    
            elif animal_str in kit2_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                else:
                    kit2_nevidno = False
                    
            elif animal_str in kit3_aliases:
                if action_str == "nevidno":
                    kit3_nevidno = True
                else:
                    kit3_nevidno = False
                    
            elif animal_str in kit12_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit2_nevidno = True
                else:
                    kit1_nevidno = False
                    kit2_nevidno = False
                    
            elif animal_str in kit13_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit3_nevidno = True
                else:
                    kit1_nevidno = False
                    kit3_nevidno = False
                    
            elif animal_str in kit23_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                    kit3_nevidno = True
                else:
                    kit2_nevidno = False
                    kit3_nevidno = False
            
            if kit1_nevidno and kit2_nevidno and kit3_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    duration = current_time - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
        
        if simultaneous_nevidno_start_time is not None and last_time is not None:
            duration = last_time - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        
        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время одновременного невидно для 3 котят: {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")

#         LITTER OF 4 KITTEN
    elif kit_number == 4:
        print("выводок из 4 котят")
    
        output_cell = "S2"
        
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        workbook = load_workbook(excel_file)
        sheet = workbook.active
            
        total_time = 0
        kit1_nevidno = True
        kit2_nevidno = True
        kit3_nevidno = True
        kit4_nevidno = True
        simultaneous_nevidno_start_time = 0.0
        last_time = None
        
        for row in sheet.iter_rows(min_row=7, values_only=True):
            time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
            if time_val is not None:
                try:
                    t = float(time_val)
                    if last_time is None or t > last_time:
                        last_time = t
                except ValueError:
                    continue
        
        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
            "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
            "kitten1", "kitten1", "1kitten", "1kitten",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }
        
        kit2_aliases = {
            "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
            "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
            "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
            "kitten2", "kitten2", "2kitten", "2kitten",
            "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
            "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
            "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
            "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
            "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
        }
        
        kit3_aliases = {
            "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
            "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
            "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
            "kitten3", "kitten3", "3kitten", "3kitten",
            "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
            "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
            "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
            "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
            "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
        }
        
        kit4_aliases = {
            "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
            "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
            "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
            "kitten4", "kitten4", "4kitten", "4kitten",
            "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
            "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
            "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
            "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
            "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
        }
    
        kit12_aliases = {"12", "21", "1221", "2112"}
        kit13_aliases = {"13", "31", "1331", "3113"}
        kit14_aliases = {"14", "41", "1441", "4114"}
        kit23_aliases = {"23", "32", "2332", "3223"}
        kit24_aliases = {"24", "42", "2442", "4224"}
        kit34_aliases = {"34", "43", "3443", "4334"}
        
        kit123_aliases = {"123", "132", "213", "231", "312", "321"}
        kit124_aliases = {"124", "142", "214", "241", "412", "421"}
        kit134_aliases = {"134", "143", "314", "341", "413", "431"}
        kit234_aliases = {"234", "243", "324", "342", "423", "432"}
        
        kit1234_aliases = {"1234", "1243", "1324", "1342", "1423", "1432", "2134", "2143", "2314", "2341", "2413", "2431",
                           "3124", "3142", "3214", "3241", "3412", "3421", "4123", "4132", "4213", "4231", "4312", "4321"}
        
        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
            animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if action is None or animal is None or time is None:
                continue

            action_str = str(action).strip().lower()
            animal_str = str(animal).strip().lower()
            
            try:
                current_time = float(time)
            except ValueError:
                continue
            
            if animal_str in kit1_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                else:
                    kit1_nevidno = False
                    
            elif animal_str in kit2_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                else:
                    kit2_nevidno = False
                    
            elif animal_str in kit3_aliases:
                if action_str == "nevidno":
                    kit3_nevidno = True
                else:
                    kit3_nevidno = False
                    
            elif animal_str in kit4_aliases:
                if action_str == "nevidno":
                    kit4_nevidno = True
                else:
                    kit4_nevidno = False
                
            elif animal_str in kit12_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit2_nevidno = True
                else:
                    kit1_nevidno = False
                    kit2_nevidno = False
                    
            elif animal_str in kit13_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit3_nevidno = True
                else:
                    kit1_nevidno = False
                    kit3_nevidno = False
                    
            elif animal_str in kit14_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit4_nevidno = True
                else:
                    kit1_nevidno = False
                    kit4_nevidno = False
                    
            elif animal_str in kit23_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                    kit3_nevidno = True
                else:
                    kit2_nevidno = False
                    kit3_nevidno = False
                    
            elif animal_str in kit24_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                    kit4_nevidno = True
                else:
                    kit2_nevidno = False
                    kit4_nevidno = False
                    
            elif animal_str in kit34_aliases:
                if action_str == "nevidno":
                    kit3_nevidno = True
                    kit4_nevidno = True
                else:
                    kit3_nevidno = False
                    kit4_nevidno = False
                    
            elif animal_str in kit123_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit2_nevidno = True
                    kit3_nevidno = True
                else:
                    kit1_nevidno = False
                    kit2_nevidno = False
                    kit3_nevidno = False
                    
            elif animal_str in kit124_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit2_nevidno = True
                    kit4_nevidno = True
                else:
                    kit1_nevidno = False
                    kit2_nevidno = False
                    kit4_nevidno = False
                    
            elif animal_str in kit134_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit3_nevidno = True
                    kit4_nevidno = True
                else:
                    kit1_nevidno = False
                    kit3_nevidno = False
                    kit4_nevidno = False
                    
            elif animal_str in kit234_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                    kit3_nevidno = True
                    kit4_nevidno = True
                else:
                    kit2_nevidno = False
                    kit3_nevidno = False
                    kit4_nevidno = False
                    
            elif animal_str in kit1234_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit2_nevidno = True
                    kit3_nevidno = True
                    kit4_nevidno = True
                else:
                    kit1_nevidno = False
                    kit2_nevidno = False
                    kit3_nevidno = False
                    kit4_nevidno = False
            
            if kit1_nevidno and kit2_nevidno and kit3_nevidno and kit4_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    duration = current_time - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
        
        if simultaneous_nevidno_start_time is not None and last_time is not None:
            duration = last_time - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        
        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время одновременного невидно для 4 котят: {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")



#         LITTER OF 5 KITTEN
    elif kit_number == 5:
        print("выводок из 5 котят")
        
        output_cell = "S2"
        
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        total_time = 0
        kit1_nevidno = True
        kit2_nevidno = True
        kit3_nevidno = True
        kit4_nevidno = True
        kit5_nevidno = True
        
        simultaneous_nevidno_start_time = 0.0
        last_time = None
        
        for row in sheet.iter_rows(min_row=7, values_only=True):
            time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
            if time_val is not None:
                try:
                    t = float(time_val)
                    if last_time is None or t > last_time:
                        last_time = t
                except ValueError:
                    continue
    
        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
            "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
            "kitten1", "kitten1", "1kitten", "1kitten",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }
        
        kit2_aliases = {
            "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
            "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
            "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
            "kitten2", "kitten2", "2kitten", "2kitten",
            "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
            "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
            "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
            "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
            "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
        }
        
        kit3_aliases = {
            "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
            "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
            "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
            "kitten3", "kitten3", "3kitten", "3kitten",
            "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
            "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
            "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
            "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
            "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
        }
        
        kit4_aliases = {
            "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
            "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
            "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
            "kitten4", "kitten4", "4kitten", "4kitten",
            "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
            "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
            "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
            "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
            "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
        }
        
        kit5_aliases = {
            "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
            "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
            "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
            "kitten5", "kitten5", "5kitten", "5kitten",
            "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
            "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
            "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
            "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
            "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
        }
        
        kit12_aliases = {"12", "21", "1221", "2112"}
        kit13_aliases = {"13", "31", "1331", "3113"}
        kit14_aliases = {"14", "41", "1441", "4114"}
        kit15_aliases = {"15", "51", "1551", "5115"}
        kit23_aliases = {"23", "32", "2332", "3223"}
        kit24_aliases = {"24", "42", "2442", "4224"}
        kit25_aliases = {"25", "52", "2552", "5225"}
        kit34_aliases = {"34", "43", "3443", "4334"}
        kit35_aliases = {"35", "53", "3553", "5335"}
        kit45_aliases = {"45", "54", "4554", "5445"}
    
        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
            animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if action is None or animal is None or time is None:
                continue

            action_str = str(action).strip().lower()
            animal_str = str(animal).strip().lower()
            
            try:
                current_time = float(time)
            except ValueError:
                continue
            
            if animal_str in kit1_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                else:
                    kit1_nevidno = False
                
            elif animal_str in kit2_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                else:
                    kit2_nevidno = False
                    
            elif animal_str in kit3_aliases:
                if action_str == "nevidno":
                    kit3_nevidno = True
                else:
                    kit3_nevidno = False
                    
            elif animal_str in kit4_aliases:
                if action_str == "nevidno":
                    kit4_nevidno = True
                else:
                    kit4_nevidno = False
                    
            elif animal_str in kit5_aliases:
                if action_str == "nevidno":
                    kit5_nevidno = True
                else:
                    kit5_nevidno = False
                
            elif animal_str in kit12_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit2_nevidno = True
                else:
                    kit1_nevidno = False
                    kit2_nevidno = False
                    
            elif animal_str in kit13_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit3_nevidno = True
                else:
                    kit1_nevidno = False
                    kit3_nevidno = False
                    
            elif animal_str in kit14_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit4_nevidno = True
                else:
                    kit1_nevidno = False
                    kit4_nevidno = False
                    
            elif animal_str in kit15_aliases:
                if action_str == "nevidno":
                    kit1_nevidno = True
                    kit5_nevidno = True
                else:
                    kit1_nevidno = False
                    kit5_nevidno = False
                    
            elif animal_str in kit23_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                    kit3_nevidno = True
                else:
                    kit2_nevidno = False
                    kit3_nevidno = False
                    
            elif animal_str in kit24_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                    kit4_nevidno = True
                else:
                    kit2_nevidno = False
                    kit4_nevidno = False
                    
            elif animal_str in kit25_aliases:
                if action_str == "nevidno":
                    kit2_nevidno = True
                    kit5_nevidno = True
                else:
                    kit2_nevidno = False
                    kit5_nevidno = False
                    
            elif animal_str in kit34_aliases:
                if action_str == "nevidno":
                    kit3_nevidno = True
                    kit4_nevidno = True
                else:
                    kit3_nevidno = False
                    kit4_nevidno = False
                    
            elif animal_str in kit35_aliases:
                if action_str == "nevidno":
                    kit3_nevidno = True
                    kit5_nevidno = True
                else:
                    kit3_nevidno = False
                    kit5_nevidno = False
                    
            elif animal_str in kit45_aliases:
                if action_str == "nevidno":
                    kit4_nevidno = True
                    kit5_nevidno = True
                else:
                    kit4_nevidno = False
                    kit5_nevidno = False
            
            if kit1_nevidno and kit2_nevidno and kit3_nevidno and kit4_nevidno and kit5_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    duration = current_time - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
        
        if simultaneous_nevidno_start_time is not None and last_time is not None:
            duration = last_time - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        
        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время одновременного невидно для 5 котят: {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")



    output_cell = "R2"

    print(f"Активный лист: {sheet.title}")

    last_value = None
    for row in sheet.iter_rows(min_row=7, min_col=6, max_col=6, values_only=True):
        if row[0] is not None:
            last_value = row[0]

    if last_value is not None:
        sheet[output_cell] = last_value
        workbook.save(excel_file)
        print(f"Значение TotlTime записано в таблицу невидно.")
        
        


    workbook = load_workbook(excel_file)
    sheet = workbook.active
    
    cell_tt = "R2"
    cell_nv = "S2"
    output_cell = "T2"

    value_tt = float(sheet[cell_tt].value)
    value_nv = float(sheet[cell_nv].value)

    observed = value_tt - value_nv

    sheet[output_cell] = observed
    workbook.save(excel_file)

    print(f"Значение observed: {observed}, записана в таблицу невидно.")


    action_column = 'D'
    animal_column = 'K'

    content_to_count = "cocanie   "

    countMS = 0
    countKS = 0
    countST = 0
    countCS = 0
    countCUS = 0

    count1SS = 0
    count2SS = 0
    count3SS = 0
    count4SS = 0
    countNSS = 0

    count1USS = 0
    count2USS = 0
    count3USS = 0
    count4USS = 0
    countNUSS = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "1 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value.value) == "1 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value.value) == "1 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value.value) == "1 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value.value) == "1 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
                
            elif str(cell_additional_value.value) == "2 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value.value) == "2 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value.value) == "2 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value.value) == "2 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value.value) == "2 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
                
            elif str(cell_additional_value.value) == "3 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value.value) == "3 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value.value) == "3 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value.value) == "3 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value.value) == "3 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
            
            elif str(cell_additional_value.value) == "4 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value.value) == "4 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value.value) == "4 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value.value) == "4 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value.value) == "4 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
                
            elif str(cell_additional_value.value) == "5 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value.value) == "5 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value.value) == "5 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value.value) == "5 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value.value) == "5 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
                
            elif str(cell_additional_value.value) == "1 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value.value) == "1 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value.value) == "1 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value.value) == "1 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value.value) == "1 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
            
            elif str(cell_additional_value.value) == "2 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value.value) == "2 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value.value) == "2 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value.value) == "2 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value.value) == "2 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
                
            elif str(cell_additional_value.value) == "3 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value.value) == "3 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value.value) == "3 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value.value) == "3 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value.value) == "3 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
                
            elif str(cell_additional_value.value) == "4 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value.value) == "4 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value.value) == "4 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value.value) == "4 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value.value) == "4 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
                
            elif str(cell_additional_value.value) == "5 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value.value) == "5 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value.value) == "5 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value.value) == "5 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value.value) == "5 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
                
            elif str(cell_additional_value.value) == "1 1pairend":
                countST += 1
            elif str(cell_additional_value.value) == "1 2pairend":
                countST += 1
            elif str(cell_additional_value.value) == "1 3pairend":
                countST += 1
            elif str(cell_additional_value.value) == "1 4pairend":
                countST += 1
            elif str(cell_additional_value.value) == "1 ?pairend":
                countST += 1
                
            elif str(cell_additional_value.value) == "2 1pairend":
                countST += 1
            elif str(cell_additional_value.value) == "2 2pairend":
                countST += 1
            elif str(cell_additional_value.value) == "2 3pairend":
                countST += 1
            elif str(cell_additional_value.value) == "2 4pairend":
                countST += 1
            elif str(cell_additional_value.value) == "2 ?pairend":
                countST += 1
                
            elif str(cell_additional_value.value) == "3 1pairend":
                countST += 1
            elif str(cell_additional_value.value) == "3 2pairend":
                countST += 1
            elif str(cell_additional_value.value) == "3 3pairend":
                countST += 1
            elif str(cell_additional_value.value) == "3 4pairend":
                countST += 1
            elif str(cell_additional_value.value) == "3 ?pairend":
                countST += 1
                
            elif str(cell_additional_value.value) == "4 1pairend":
                countST += 1
            elif str(cell_additional_value.value) == "4 2pairend":
                countST += 1
            elif str(cell_additional_value.value) == "4 3pairend":
                countST += 1
            elif str(cell_additional_value.value) == "4 4pairend":
                countST += 1
            elif str(cell_additional_value.value) == "4 ?pairend":
                countST += 1
                
            elif str(cell_additional_value.value) == "5 1pairend":
                countST += 1
            elif str(cell_additional_value.value) == "5 2pairend":
                countST += 1
            elif str(cell_additional_value.value) == "5 3pairend":
                countST += 1
            elif str(cell_additional_value.value) == "5 4pairend":
                countST += 1
            elif str(cell_additional_value.value) == "5 ?pairend":
                countST += 1
                

    content_to_count = "oshibka   "

    for row in sheet.iter_rows():
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index] 

        if cell_content.value == content_to_count: 
            if str(cell_additional_value.value) == "Mstop":
                countMS += 1
            elif str(cell_additional_value.value) == "Mushla":
                countMS += 1
            elif str(cell_additional_value.value) == "Mchangedposition":
                countMS += 1

    countKS = countST - countMS
                
    values_to_write = [countMS, countKS, countCS, countCUS]

    start_column = openpyxl.utils.column_index_from_string('M')
    for i, value in enumerate(values_to_write):
        sheet.cell(row=2, column=start_column + i, value=value)

    workbook.save(excel_file)
    print(f"Резултат подсчета сосания записан в верхнюю таблицу сосания.")


    cell_ms = "M2"
    cell_ks = "N2"
    output_cell = "M3"

    value_ms = float(sheet[cell_ms].value)
    value_ks = float(sheet[cell_ks].value)

    stop = value_ms + value_ks

    if stop != 0 and value_ms != 0:
        mstop = float(value_ms * 100 / stop)
        sheet[output_cell] = mstop
        workbook.save(excel_file)
    else:
        sheet[output_cell] = 0
        workbook.save(excel_file)


    output_cell = "N3"

    value_ms = float(sheet[cell_ms].value)
    value_ks = float(sheet[cell_ks].value)

    stop = value_ms + value_ks

    if stop != 0 and value_ks != 0:
        kstop = float(value_ks * 100 / stop)
        sheet[output_cell] = kstop
        workbook.save(excel_file)
    else:
        sheet[output_cell] = 0
        workbook.save(excel_file)


    print(f"Процентное соотношение окончания сосания записано в верхнюю таблицу сосания.")


    cell_ss = "O2"
    cell_uss = "P2"
    output_cell = "O3"

    value_ss = float(sheet[cell_ss].value)
    value_uss = float(sheet[cell_uss].value)

    attempt = value_ss + value_uss

    if stop != 0 and value_ss != 0:
        success = float(value_ss * 100 / attempt)
        sheet[output_cell] = success
        workbook.save(excel_file)
    else:
        sheet[output_cell] = 0
        workbook.save(excel_file)
        

    output_cell = "P3"

    value_ss = float(sheet[cell_ss].value)
    value_uss = float(sheet[cell_uss].value)

    if stop != 0 and value_uss != 0:
        unsuccess = float(value_uss * 100 / attempt)
        sheet[output_cell] = unsuccess
        workbook.save(excel_file)
    else:
        sheet[output_cell] = 0
        workbook.save(excel_file)

    print(f"Процентное соотношение удачных и неудачных попыток записано в верхнюю таблицу сосания.")


    values_to_write = [count1SS, count2SS, count3SS, count4SS, countNSS]
    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AM{start_row + i}'] = value
        
    workbook.save(excel_file)#

    print(f"Кол-во удачных попыток сосания по парам записаны в нижнюю таблицу сосания.")


    values_to_write = [count1USS, count2USS, count3USS, count4USS, countNUSS]
    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AN{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Кол-во неудачных попыток сосания по парам записаны в нижнюю таблицу сосания.")


    p1 = count1SS + count1USS
    p2 = count2SS + count2USS
    p3 = count3SS + count3USS
    p4 = count4SS + count4USS
    pN = countNSS + countNUSS

    values_to_write = [p1, p2, p3, p4, pN]
    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AO{start_row + i}'] = value
        
    workbook.save(excel_file)#

    print(f"Кол-во total попыток сосания по парам записаны в нижнюю таблицу сосания.")

#     NEVIDNO

        # 1 KITTEN
    output_cell = "O12"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'
        
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit1_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "kitten1", "1kitten", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit1_combination_aliases = {
        "14", "41", "1441", "4114",
        "15", "51", "1551", "5115",
        "13", "31", "1331", "3113",
        "12", "21", "1221", "2112"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
    
        is_kit1_event = (animal_str in kit1_aliases) or (animal_str in kit1_combination_aliases)
        
        if is_kit1_event:
            if action_str == "nevidno":
                kit1_nevidno = True
            else:
                kit1_nevidno = False
        
        if kit1_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR6"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 1: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")

    #2 KITTEN
    output_cell = "O13"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'
        
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit2_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "kitten2", "2kitten", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    kit2_combination_aliases = {
        "24", "42", "2442", "4224",
        "25", "52", "2552", "5225",
        "23", "32", "2332", "3223",
        "21", "12", "2112", "1221"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
    
        is_kit2_event = (animal_str in kit2_aliases) or (animal_str in kit2_combination_aliases)
        
        if is_kit2_event:
            if action_str == "nevidno":
                kit2_nevidno = True
            else:
                kit2_nevidno = False
        
        if kit2_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR7"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 2: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")


    # 3 KITTEN

    output_cell = "O14"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'
        
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit3_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "kitten3", "3kitten", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    kit3_combination_aliases = {
        "31", "13", "3113", "1331",
        "32", "23", "3223", "2332",
        "34", "43", "3443", "4334",
        "35", "53", "3553", "5335"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
    
        is_kit3_event = (animal_str in kit3_aliases) or (animal_str in kit3_combination_aliases)
        
        if is_kit3_event:
            if action_str == "nevidno":
                kit3_nevidno = True
            else:
                kit3_nevidno = False
        
        if kit3_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR8"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 3: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")




    # 4 KITTEN

    output_cell = "O15"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'
        
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit4_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "kitten4", "4kitten", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    kit4_combination_aliases = {
        "41", "14", "4114", "1441",
        "42", "24", "4224", "2442",
        "34", "43", "3443", "4334",
        "45", "54", "4554", "5445"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
    
        is_kit4_event = (animal_str in kit4_aliases) or (animal_str in kit4_combination_aliases)
        
        if is_kit4_event:
            if action_str == "nevidno":
                kit4_nevidno = True
            else:
                kit4_nevidno = False
        
        if kit4_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR9"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 4: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")




    # 5 KITTEN

    output_cell = "O16"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'
        
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "kitten5", "5kitten", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    kit5_combination_aliases = {
        "51", "15", "5115", "1551",
        "52", "25", "5225", "2552",
        "53", "35", "5335", "3553",
        "54", "45", "5445", "4554"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
    
        is_kit5_event = (animal_str in kit5_aliases) or (animal_str in kit5_combination_aliases)
        
        if is_kit5_event:
            if action_str == "nevidno":
                kit5_nevidno = True
            else:
                kit5_nevidno = False
        
        if kit5_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR10"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 5: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # MOTHER

    output_cell = "O11"
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    total_time = 0
    kitM_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kitM_aliases = {
        "m", "mm", "1m", "m1", "1mm1", "m11m", 
        "mneighbour", "neighbourm", "mneighbourneighbourm", "neighbourmmneighbour",
        "mneighbor", "neighborm", "mneighborneighborm", "neighbormmneighbor",
        "2m", "m2", "2mm2", "m22m",
        "3m", "m3", "3mm3", "m33m",
        "4m", "m4", "4mm4", "m44m",
        "5m", "m5", "5mm5", "m55m",
        "Kittenm", "kittenm", "mKitten", "mkitten"
    }


    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kitM_event = animal_str in kitM_aliases
        
        if action_str == "nevidno":
            if is_kitM_event:
                kitM_nevidno = True
        elif action_str == "vokal":
            continue
        elif action_str == "cocanie":
            kitM_nevidno = False
        else:
            if is_kitM_event:
                kitM_nevidno = False
        
        if kitM_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR11"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для мамы: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 1 END 2

    output_cell = "O27"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit1_nevidno = True
    kit2_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "kitten1", "1kitten", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "kitten2", "2kitten", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    both_kits_aliases = {
        "12", "21", "1221", "2112"
    }

    kit1_with_others_aliases = {
        "14", "41", "1441", "4114",
        "15", "51", "1551", "5115",
        "13", "31", "1331", "3113"
    }

    kit2_with_others_aliases = {
        "24", "42", "2442", "4224",
        "23", "32", "2332", "3223",
        "25", "52", "2552", "5225"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit1 = animal_str in kit1_aliases
        is_kit2 = animal_str in kit2_aliases
        is_both = animal_str in both_kits_aliases
        is_kit1_with_other = animal_str in kit1_with_others_aliases
        is_kit2_with_other = animal_str in kit2_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit1:
                kit1_nevidno = True
            if is_kit2:
                kit2_nevidno = True
            if is_both:
                kit1_nevidno = True
                kit2_nevidno = True
            if is_kit1_with_other:
                kit1_nevidno = True
            if is_kit2_with_other:
                kit2_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit1:
                kit1_nevidno = False
            if is_kit2:
                kit2_nevidno = False
            if is_both:
                kit1_nevidno = False
                kit2_nevidno = False
            if is_kit1_with_other:
                kit1_nevidno = False
            if is_kit2_with_other:
                kit2_nevidno = False
        
        if kit1_nevidno and kit2_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR12"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 1 и 2: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")




    # 1 END 3

    output_cell = "O28"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit1_nevidno = True
    kit3_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "kitten1", "1kitten", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "kitten3", "3kitten", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    both_kits_aliases = {
        "13", "31", "1331", "3113"
    }

    kit1_with_others_aliases = {
        "14", "41", "1441", "4114",
        "15", "51", "1551", "5115",
        "12", "21", "1221", "2112"
    }

    kit3_with_others_aliases = {
        "34", "43", "3443", "4334",
        "23", "32", "2332", "3223",
        "35", "53", "3553", "5335"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit1 = animal_str in kit1_aliases
        is_kit3 = animal_str in kit3_aliases
        is_both = animal_str in both_kits_aliases
        is_kit1_with_other = animal_str in kit1_with_others_aliases
        is_kit3_with_other = animal_str in kit3_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit1:
                kit1_nevidno = True
            if is_kit3:
                kit3_nevidno = True
            if is_both:
                kit1_nevidno = True
                kit3_nevidno = True
            if is_kit1_with_other:
                kit1_nevidno = True
            if is_kit3_with_other:
                kit3_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit1:
                kit1_nevidno = False
            if is_kit3:
                kit3_nevidno = False
            if is_both:
                kit1_nevidno = False
                kit3_nevidno = False
            if is_kit1_with_other:
                kit1_nevidno = False
            if is_kit3_with_other:
                kit3_nevidno = False
        
        if kit1_nevidno and kit3_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR13"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 1 и 3: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 1 END 4

    output_cell = "O29"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit1_nevidno = True
    kit4_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "kitten1", "1kitten", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "kitten4", "4kitten", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    both_kits_aliases = {
        "14", "41", "1441", "4114"
    }

    kit1_with_others_aliases = {
        "15", "51", "1551", "5115",
        "13", "31", "1331", "3113",
        "12", "21", "1221", "2112"
    }

    kit4_with_others_aliases = {
        "24", "42", "2442", "4224",
        "34", "43", "3443", "4334",
        "45", "54", "4554", "5445"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit1 = animal_str in kit1_aliases
        is_kit4 = animal_str in kit4_aliases
        is_both = animal_str in both_kits_aliases
        is_kit1_with_other = animal_str in kit1_with_others_aliases
        is_kit4_with_other = animal_str in kit4_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit1:
                kit1_nevidno = True
            if is_kit4:
                kit4_nevidno = True
            if is_both:
                kit1_nevidno = True
                kit4_nevidno = True
            if is_kit1_with_other:
                kit1_nevidno = True
            if is_kit4_with_other:
                kit4_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit1:
                kit1_nevidno = False
            if is_kit4:
                kit4_nevidno = False
            if is_both:
                kit1_nevidno = False
                kit4_nevidno = False
            if is_kit1_with_other:
                kit1_nevidno = False
            if is_kit4_with_other:
                kit4_nevidno = False
        
        if kit1_nevidno and kit4_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR14"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 1 и 4: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 1 END 5

    output_cell = "O30"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit1_nevidno = True
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "kitten1", "1kitten", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "kitten5", "5kitten", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    both_kits_aliases = {
        "15", "51", "1551", "5115"
    }

    kit1_with_others_aliases = {
        "14", "41", "1441", "4114",
        "13", "31", "1331", "3113",
        "12", "21", "1221", "2112"
    }

    kit5_with_others_aliases = {
        "25", "52", "2552", "5225",
        "35", "53", "3553", "5335",
        "45", "54", "4554", "5445"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit1 = animal_str in kit1_aliases
        is_kit5 = animal_str in kit5_aliases
        is_both = animal_str in both_kits_aliases
        is_kit1_with_other = animal_str in kit1_with_others_aliases
        is_kit5_with_other = animal_str in kit5_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit1:
                kit1_nevidno = True
            if is_kit5:
                kit5_nevidno = True
            if is_both:
                kit1_nevidno = True
                kit5_nevidno = True
            if is_kit1_with_other:
                kit1_nevidno = True
            if is_kit5_with_other:
                kit5_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit1:
                kit1_nevidno = False
            if is_kit5:
                kit5_nevidno = False
            if is_both:
                kit1_nevidno = False
                kit5_nevidno = False
            if is_kit1_with_other:
                kit1_nevidno = False
            if is_kit5_with_other:
                kit5_nevidno = False
        
        if kit1_nevidno and kit5_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR15"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 1 и 5: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 2 END 3

    output_cell = "O32"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit2_nevidno = True
    kit3_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "kitten2", "2kitten", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "kitten3", "3kitten", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    both_kits_aliases = {
        "23", "32", "2332", "3223"
    }

    kit2_with_others_aliases = {
        "12", "21", "1221", "2112",
        "24", "42", "2442", "4224",
        "25", "52", "2552", "5225"
    }

    kit3_with_others_aliases = {
        "13", "31", "1331", "3113",
        "34", "43", "3443", "4334",
        "35", "53", "3553", "5335"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit2 = animal_str in kit2_aliases
        is_kit3 = animal_str in kit3_aliases
        is_both = animal_str in both_kits_aliases
        is_kit2_with_other = animal_str in kit2_with_others_aliases
        is_kit3_with_other = animal_str in kit3_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit2:
                kit2_nevidno = True
            if is_kit3:
                kit3_nevidno = True
            if is_both:
                kit2_nevidno = True
                kit3_nevidno = True
            if is_kit2_with_other:
                kit2_nevidno = True
            if is_kit3_with_other:
                kit3_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit2:
                kit2_nevidno = False
            if is_kit3:
                kit3_nevidno = False
            if is_both:
                kit2_nevidno = False
                kit3_nevidno = False
            if is_kit2_with_other:
                kit2_nevidno = False
            if is_kit3_with_other:
                kit3_nevidno = False
        
        if kit2_nevidno and kit3_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR16"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 2 и 3: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 2 END 4

    output_cell = "O33"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit2_nevidno = True
    kit4_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "kitten2", "2kitten", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "kitten4", "4kitten", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    both_kits_aliases = {
        "24", "42", "2442", "4224"
    }

    kit2_with_others_aliases = {
        "12", "21", "1221", "2112",
        "23", "32", "2332", "3223",
        "25", "52", "2552", "5225"
    }

    kit4_with_others_aliases = {
        "14", "41", "1441", "4114",
        "34", "43", "3443", "4334",
        "45", "54", "4554", "5445"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit2 = animal_str in kit2_aliases
        is_kit4 = animal_str in kit4_aliases
        is_both = animal_str in both_kits_aliases
        is_kit2_with_other = animal_str in kit2_with_others_aliases
        is_kit4_with_other = animal_str in kit4_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit2:
                kit2_nevidno = True
            if is_kit4:
                kit4_nevidno = True
            if is_both:
                kit2_nevidno = True
                kit4_nevidno = True
            if is_kit2_with_other:
                kit2_nevidno = True
            if is_kit4_with_other:
                kit4_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit2:
                kit2_nevidno = False
            if is_kit4:
                kit4_nevidno = False
            if is_both:
                kit2_nevidno = False
                kit4_nevidno = False
            if is_kit2_with_other:
                kit2_nevidno = False
            if is_kit4_with_other:
                kit4_nevidno = False
        
        if kit2_nevidno and kit4_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR17"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 2 и 4: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 2 END 5

    output_cell = "O34"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit2_nevidno = True
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "kitten2", "2kitten", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "kitten5", "5kitten", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    both_kits_aliases = {
        "25", "52", "2552", "5225"
    }

    kit2_with_others_aliases = {
        "12", "21", "1221", "2112",
        "23", "32", "2332", "3223",
        "24", "42", "2442", "4224"
    }

    kit5_with_others_aliases = {
        "15", "51", "1551", "5115",
        "35", "53", "3553", "5335",
        "45", "54", "4554", "5445"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit2 = animal_str in kit2_aliases
        is_kit5 = animal_str in kit5_aliases
        is_both = animal_str in both_kits_aliases
        is_kit2_with_other = animal_str in kit2_with_others_aliases
        is_kit5_with_other = animal_str in kit5_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit2:
                kit2_nevidno = True
            if is_kit5:
                kit5_nevidno = True
            if is_both:
                kit2_nevidno = True
                kit5_nevidno = True
            if is_kit2_with_other:
                kit2_nevidno = True
            if is_kit5_with_other:
                kit5_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit2:
                kit2_nevidno = False
            if is_kit5:
                kit5_nevidno = False
            if is_both:
                kit2_nevidno = False
                kit5_nevidno = False
            if is_kit2_with_other:
                kit2_nevidno = False
            if is_kit5_with_other:
                kit5_nevidno = False
        
        if kit2_nevidno and kit5_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR18"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 2 и 5: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 3 END 4

    output_cell = "O37"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit3_nevidno = True
    kit4_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "kitten3", "3kitten", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "kitten4", "4kitten", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    both_kits_aliases = {
        "34", "43", "3443", "4334"
    }

    kit3_with_others_aliases = {
        "13", "31", "1331", "3113",
        "23", "32", "2332", "3223",
        "35", "53", "3553", "5335"
    }

    kit4_with_others_aliases = {
        "14", "41", "1441", "4114",
        "24", "42", "2442", "4224",
        "45", "54", "4554", "5445"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit3 = animal_str in kit3_aliases
        is_kit4 = animal_str in kit4_aliases
        is_both = animal_str in both_kits_aliases
        is_kit3_with_other = animal_str in kit3_with_others_aliases
        is_kit4_with_other = animal_str in kit4_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit3:
                kit3_nevidno = True
            if is_kit4:
                kit4_nevidno = True
            if is_both:
                kit3_nevidno = True
                kit4_nevidno = True
            if is_kit3_with_other:
                kit3_nevidno = True
            if is_kit4_with_other:
                kit4_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit3:
                kit3_nevidno = False
            if is_kit4:
                kit4_nevidno = False
            if is_both:
                kit3_nevidno = False
                kit4_nevidno = False
            if is_kit3_with_other:
                kit3_nevidno = False
            if is_kit4_with_other:
                kit4_nevidno = False
        
        if kit3_nevidno and kit4_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR19"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 3 и 4: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 3 END 5

    output_cell = "O38"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit3_nevidno = True
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "kitten3", "3kitten", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "kitten5", "5kitten", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    both_kits_aliases = {
        "35", "53", "3553", "5335"
    }

    kit3_with_others_aliases = {
        "13", "31", "1331", "3113",
        "23", "32", "2332", "3223",
        "34", "43", "3443", "4334"
    }

    kit5_with_others_aliases = {
        "15", "51", "1551", "5115",
        "25", "52", "2552", "5225",
        "45", "54", "4554", "5445"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit3 = animal_str in kit3_aliases
        is_kit5 = animal_str in kit5_aliases
        is_both = animal_str in both_kits_aliases
        is_kit3_with_other = animal_str in kit3_with_others_aliases
        is_kit5_with_other = animal_str in kit5_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit3:
                kit3_nevidno = True
            if is_kit5:
                kit5_nevidno = True
            if is_both:
                kit3_nevidno = True
                kit5_nevidno = True
            if is_kit3_with_other:
                kit3_nevidno = True
            if is_kit5_with_other:
                kit5_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit3:
                kit3_nevidno = False
            if is_kit5:
                kit5_nevidno = False
            if is_both:
                kit3_nevidno = False
                kit5_nevidno = False
            if is_kit3_with_other:
                kit3_nevidno = False
            if is_kit5_with_other:
                kit5_nevidno = False
        
        if kit3_nevidno and kit5_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR20"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 3 и 5: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 4 END 5

    output_cell = "O42"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    total_time = 0
    kit4_nevidno = True
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_time = None

    for row in sheet.iter_rows(min_row=7, values_only=True):
        time_val = row[openpyxl.utils.column_index_from_string(time_column) - 1]
        if time_val is not None:
            try:
                t = float(time_val)
                if last_time is None or t > last_time:
                    last_time = t
            except ValueError:
                continue

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "kitten4", "4kitten", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "kitten5", "5kitten", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    both_kits_aliases = {
        "45", "54", "4554", "5445"
    }

    kit4_with_others_aliases = {
        "14", "41", "1441", "4114",
        "24", "42", "2442", "4224",
        "34", "43", "3443", "4334"
    }

    kit5_with_others_aliases = {
        "15", "51", "1551", "5115",
        "25", "52", "2552", "5225",
        "35", "53", "3553", "5335"
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        action = row[openpyxl.utils.column_index_from_string(action_column) - 1]
        animal = row[openpyxl.utils.column_index_from_string(animal_column) - 1]
        time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

        if action is None or animal is None or time is None:
            continue

        action_str = str(action).strip().lower()
        animal_str = str(animal).strip().lower()
        
        try:
            current_time = float(time)
        except ValueError:
            continue
        
        is_kit4 = animal_str in kit4_aliases
        is_kit5 = animal_str in kit5_aliases
        is_both = animal_str in both_kits_aliases
        is_kit4_with_other = animal_str in kit4_with_others_aliases
        is_kit5_with_other = animal_str in kit5_with_others_aliases
        
        if action_str == "nevidno":
            if is_kit4:
                kit4_nevidno = True
            if is_kit5:
                kit5_nevidno = True
            if is_both:
                kit4_nevidno = True
                kit5_nevidno = True
            if is_kit4_with_other:
                kit4_nevidno = True
            if is_kit5_with_other:
                kit5_nevidno = True
                
        elif action_str == "vokal":
            continue
            
        else:
            if is_kit4:
                kit4_nevidno = False
            if is_kit5:
                kit5_nevidno = False
            if is_both:
                kit4_nevidno = False
                kit5_nevidno = False
            if is_kit4_with_other:
                kit4_nevidno = False
            if is_kit5_with_other:
                kit5_nevidno = False
        
        if kit4_nevidno and kit5_nevidno:
            if simultaneous_nevidno_start_time is None:
                simultaneous_nevidno_start_time = current_time
        else:
            if simultaneous_nevidno_start_time is not None:
                duration = current_time - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
                simultaneous_nevidno_start_time = None

    if simultaneous_nevidno_start_time is not None and last_time is not None:
        duration = last_time - simultaneous_nevidno_start_time
        if duration > 0:
            total_time += duration

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR21"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 4 и 5: {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")







    def analyze_animal_play(excel_file):
        try:
            workbook = load_workbook(excel_file)  # Открываем файл для чтения и записи
            sheet = workbook.active  # Получаем активный лист
        except FileNotFoundError:
            print(f"Ошибка: Файл {excel_file} не найден.")
            return
        except Exception as e:
            print(f"Ошибка при открытии файла: {e}")
            return

        games = []
        current_borba = None
        start_row = 7

        animal_pairs_duration = {}

        for index, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row - 1):  # Итерируемся, начиная с нужной строки
            try:
                action = str(row[3].value).strip().lower() if row[3].value is not None else ''
                animals = str(row[10].value).strip() if row[10].value is not None else ''  # Убрал lower() чтобы сохранить оригинальный регистр
                time = row[5].value

                if 'borba' in action.lower() and not animals.lower().endswith('end'):
                    if current_borba:
                        print(f"Борьба {current_borba['animals']} начата в строке {current_borba['start_row']}, но не завершена.")
                    current_borba = {
                        'start_time': time,
                        'animals': animals,
                        'start_row': index + 1
                    }

                elif 'borba' in action.lower() and animals.lower().endswith('end'):
                    if current_borba:
                        # Сравниваем без учета регистра и удаляем 'end' (3 символа)
                        if current_borba['animals'].lower() == animals[:-3].lower():
                            current_borba['end_time'] = time
                            current_borba['end_row'] = index + 1

                            try:
                                start_time = float(current_borba['start_time'])
                                end_time = float(current_borba['end_time'])
                                current_borba['duration'] = round(end_time - start_time, 2)
                                pair_key = current_borba['animals']  # Сохраняем оригинальное значение (например, '1M')
                                if pair_key in animal_pairs_duration:
                                    animal_pairs_duration[pair_key] += current_borba['duration']
                                else:
                                    animal_pairs_duration[pair_key] = current_borba['duration']

                            except (ValueError, TypeError) as e:
                                print(f"Ошибка преобразования времени в строке {index + 1}: {e}")
                                current_borba['duration'] = None
    
                            games.append(current_borba)
                            current_borba = None

                        else:
                            print(f"Несоответствие участников борьбы. Ожидалось {current_borba['animals']}, получено {animals}.")
                    else:
                        print(f"Обнаружено окончание борьбы без начала в строке {index + 1}.")

            except Exception as e:
                print(f"Ошибка в строке {index + 1}: {e}, Значение animals: {row[10].value}, Тип данных animals: {type(row[10].value)}")

        print(f"\nКоличество борьбы: {len(games)}")
        for i, borba in enumerate(games, 1):
            print(f"\nБорьба {i}:")
            print(f"Участники: {borba['animals']}")
            print(f"Начало борьбы: строка {borba['start_row']}, время {borba['start_time']} сек")
            print(f"Окончание борьбы: строка {borba['end_row']}, время {borba['end_time']} сек")
            print(f"Продолжительность борьбы: {borba['duration']} сек")

        output_row = 6
        for pair, duration in animal_pairs_duration.items():
            sheet[f'AG{output_row}'] = pair
            sheet[f'AH{output_row}'] = duration
            output_row += 1

        try:
            workbook.save(excel_file)
            print(f"\nДанные по группам борьбы записаны в столбцы AG и AH.")
        except Exception as e:
            print(f"Ошибка при сохранении файла: {e}")

    def analyze_animal_groups(excel_file):
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
        except FileNotFoundError:
            print(f"Ошибка: Файл {excel_file} не найден.")
            return
        except Exception as e:
            print(f"Ошибка при открытии файла: {e}")
            return

        total_duration = 0
        group1_duration = 0
        group2_duration = 0
        group3_duration = 0

        start_row = 6

        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, min_col=33, max_col=34, values_only=True), start=start_row):
            animals = str(row[0]) if row[0] is not None else ''
            duration_value = row[1]

            try:
                duration = float(duration_value) if duration_value is not None else 0
            except (ValueError, TypeError):
                print(f"Ошибка: Невозможно преобразовать значение '{duration_value}' в число в строке {row_idx}. Пропускаем эту строку.")
                continue

            total_duration += duration
            
            if animals.isdigit():
                group1_duration += duration
            elif animals.lower().startswith('m'):
                group2_duration += duration
            elif animals.lower().endswith('m'):
                group3_duration += duration

        sheet['AJ6'] = total_duration
        sheet['AJ7'] = group1_duration
        sheet['AJ8'] = group2_duration
        sheet['AJ9'] = group3_duration

        try:
            workbook.save(excel_file)
            print("Результаты записаны в столбец AJ.")
        except Exception as e:
            print(f"Ошибка при сохранении файла: {e}")

    analyze_animal_play(excel_file)
    analyze_animal_groups(excel_file)






#     2 PROTOCOL
elif protocol_number == 2:
    print("ПРОТОКОЛ 2")
    
#     ENTRY END OF 2 MINUTES
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook.active

    time_column = 'F'

    last_time_cell = None

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_index in range(7, sheet.max_row + 1):
        current_time = sheet[f"{time_column}{row_index}"].value

        if current_time is None:
            continue

        current_time = float(current_time)

        if current_time < end_time:
            last_time_cell = row_index

    if last_time_cell:
        end_time_cell = f"L{last_time_cell}"
        sheet[end_time_cell] = end_time

        sheet[end_time_cell].fill = yellow_fill

        print(f"Значение времени конца учета данных записано в ячейку {end_time_cell}.")
    else:
        print("Следует учесть все данные в файле.")

    workbook.save(excel_file)
    
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    content_to_count = "allogrumin"

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                    
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'P{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета аллогруминга записан в сотлбец P.")


    content_to_count = "gruming   "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                 
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'Q{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета груминга записан в сотлбец Q.")


    content_to_count = "igra      "


    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
            
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'R{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета игры записан в сотлбец R.")


    content_to_count = "igra      "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0


    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "11":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "22":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "33":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "44":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "55":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "MM":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'S{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета selfplay записан в сотлбец S.")


    content_to_count = "igrasmamoj"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
                         
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'T{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета игра с активной мамой записан в сотлбец T.")


    content_to_count = "igrasmamoj"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "1tail":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "1ears":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "1paw":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2ears":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "2tail":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "2paw":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3ears":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "3paw":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "3tail":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4ears":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "4tail":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "4paw":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5ears":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "5paw":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "5tail":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittentail":
                countKM += 1
            elif str(cell_additional_value) == "Kittentail":
                countKM += 1
            elif str(cell_additional_value) == "kittenears":
                countKM += 1
            elif str(cell_additional_value) == "Kittenears":
                countKM += 1
            elif str(cell_additional_value) == "kittenpaw":
                countKM += 1
            elif str(cell_additional_value) == "Kittenpaw":
                countKM += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'U{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета игра с пассивной мамой записан в сотлбец U.")


    content_to_count = "spredmetom"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                 
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'V{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета игра с предметом записан в сотлбец V.")

     

    content_to_count = "bokom     "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
            
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                 
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'W{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета боком записан в сотлбец W.")


    content_to_count = "ckradivan "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                                    
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'X{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета обхват записан в сотлбец X.")


    content_to_count = "zataivanie"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                                    
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'Y{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета затаивание записан в сотлбец Y.")


    content_to_count = "lapki     "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0


    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M11M":
                countMK += 1
                countM1 += 1
                count1M += 1
                countKM += 1
            elif str(cell_additional_value) == "M22M":
                countMK += 1
                countM2 += 1
                count2M += 1
                countKM += 1
            elif str(cell_additional_value) == "M33M":
                countMK += 1
                countM3 += 1
                count3M += 1
                countKM += 1
            elif str(cell_additional_value) == "M44M":
                countMK += 1
                countM4 += 1
                count4M += 1
                countKM += 1
            elif str(cell_additional_value) == "M55M":
                countMK += 1
                countM5 += 1
                count5M += 1
                countKM += 1
            elif str(cell_additional_value) == "MkittenkittemM":
                countMK += 1
                countKM += 1
            elif str(cell_additional_value) == "MKittenKittenM":
                countMK += 1
                countKM += 1
            elif str(cell_additional_value) == "MKittenkittenM":
                countMK += 1
                countKM += 1
            
            
            elif str(cell_additional_value) == "1MM1":
                countMK += 1
                countM1 += 1
                count1M += 1
                countKM += 1
            elif str(cell_additional_value) == "2MM2":
                countMK += 1
                countM2 += 1
                count2M += 1
                countKM += 1
            elif str(cell_additional_value) == "3MM3":
                countMK += 1
                countM3 += 1
                count3M += 1
                countKM += 1
            elif str(cell_additional_value) == "4MM4":
                countMK += 1
                countM4 += 1
                count4M += 1
                countKM += 1
            elif str(cell_additional_value) == "5MM5":
                countMK += 1
                countM5 += 1
                count5M += 1
                countKM += 1
            elif str(cell_additional_value) == "kittenMMkitten":
                countMK += 1
                countKM += 1
            elif str(cell_additional_value) == "KittenMMKitten":
                countMK += 1
                countKM += 1
            elif str(cell_additional_value) == "KittenMMkitten":
                countMK += 1
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbourneighbour1":
                count1N += 1
                countKN += 1
                countN1 += 1
                countNK += 1
            elif str(cell_additional_value) == "2neighbourneighbour2":
                count2N += 1
                countKN += 1
                countN2 += 1
                countNK += 1
            elif str(cell_additional_value) == "3neighbourneighbour3":
                count3N += 1
                countKN += 1
                countN3 += 1
                countNK += 1
            elif str(cell_additional_value) == "4neighbourneighbour4":
                count4N += 1
                countKN += 1
                countN4 += 1
                countNK += 1
            elif str(cell_additional_value) == "5neighbourneighbour5":
                count5N += 1
                countKN += 1
                countN5 += 1
                countNK += 1
            elif str(cell_additional_value) == "MneighbourneighbourM":
                countMN += 1
                countNM += 1
            elif str(cell_additional_value) == "KittenneighbourneighbourKitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "kittenneighbourneighbourkitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "Kittenneighbourneighbourkitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "kittenneighbourneighbourKitten":
                countKN += 1
                countNK += 1
                
            elif str(cell_additional_value) == "neighbour11neighbour":
                count1N += 1
                countKN += 1
                countN1 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbour22neighbour":
                count2N += 1
                countKN += 1
                countN2 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbour33neighbour":
                count3N += 1
                countKN += 1
                countN3 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbour44neighbour":
                count4N += 1
                countKN += 1
                countN4 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbour55neighbour":
                count5N += 1
                countKN += 1
                countN5 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbourMMneighbour":
                countNM += 1
                countMN += 1
            elif str(cell_additional_value) == "neighbourKittenKittenneighbour":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbourkittenkittenneighbour":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbourKittenkittenneighbour":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbourkittenKittenneighbour":
                countKN += 1
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighborneighbor1":
                count1N += 1
                countKN += 1
                countN1 += 1
                countNK += 1
            elif str(cell_additional_value) == "2neighborneighbor2":
                count2N += 1
                countKN += 1
                countN2 += 1
                countNK += 1
            elif str(cell_additional_value) == "3neighborneighbor3":
                count3N += 1
                countKN += 1
                countN3 += 1
                countNK += 1
            elif str(cell_additional_value) == "4neighborneighbor4":
                count4N += 1
                countKN += 1
                countN4 += 1
                countNK += 1
            elif str(cell_additional_value) == "5neighborneighbor5":
                count5N += 1
                countKN += 1
                countN5 += 1
                countNK += 1
            elif str(cell_additional_value) == "MneighborneighborM":
                countMN += 1
                countNM += 1
            elif str(cell_additional_value) == "KittenneighborneighborKitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "kittenneighborneighborkitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "Kittenneighborneighborkitten":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "kittenneighborneighborKitten":
                countKN += 1
                countNK += 1
                
            elif str(cell_additional_value) == "neighbor11neighbor":
                count1N += 1
                countKN += 1
                countN1 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbor22neighbor":
                count2N += 1
                countKN += 1
                countN2 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbor33neighbor":
                count3N += 1
                countKN += 1
                countN3 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbor44neighbor":
                count4N += 1
                countKN += 1
                countN4 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighbor55neighbor":
                count5N += 1
                countKN += 1
                countN5 += 1
                countNK += 1
            elif str(cell_additional_value) == "neighborMMneighbor":
                countNM += 1
                countMN += 1
            elif str(cell_additional_value) == "neighborKittenKittenneighbor":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "neighborkittenkittenneighbor":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "neighborKittenkittenneighbor":
                countKN += 1
                countNK += 1
            elif str(cell_additional_value) == "neighborkittenKittenneighbor":
                countKN += 1
                countNK += 1
                      
            elif str(cell_additional_value) == "1221":
                countKK += 1
                count12 += 1
                count21 += 1
                count1init += 1
                count1rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value) == "1331":
                countKK += 1
                count13 += 1
                count31 += 1
                count1init += 1
                count1rec += 1
                count3init += 1
                count3rec += 1
            elif str(cell_additional_value) == "1441":
                countKK += 1
                count14 += 1
                count41 += 1
                count1init += 1
                count1rec += 1
                count4init += 1
                count4rec += 1
            elif str (cell_additional_value) == "1551":
                countKK += 1
                count15 += 1
                count51 += 1
                count1init += 1
                count1rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value) == "2112":
                countKK += 1
                count12 += 1
                count21 += 1
                count1init += 1
                count1rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value) == "2332":
                countKK += 1
                count32 += 1
                count23 += 1
                count3init += 1
                count3rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value) == "2442":
                countKK += 1
                count42 += 1
                count24 += 1
                count4init += 1
                count4rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value) == "2552":
                countKK += 1
                count52 += 1
                count25 += 1
                count5init += 1
                count5rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value) == "3113":
                countKK += 1
                count31 += 1
                count13 += 1
                count1init += 1
                count1rec += 1
                count3init += 1
                count3rec += 1
            elif str(cell_additional_value) == "3223":
                countKK += 1
                count32 += 1
                count23 += 1
                count3init += 1
                count3rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value) == "3443":
                countKK += 1
                count34 += 1
                count43 += 1
                count3init += 1
                count3rec += 1
                count4init += 1
                count4rec += 1
            elif str(cell_additional_value) == "3553":
                countKK += 1
                count35 += 1
                count53 += 1
                count3init += 1
                count3rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value) == "4114":
                countKK += 1
                count41 += 1
                count14 += 1
                count1init += 1
                count1rec += 1
                count4init += 1
                count4rec += 1
            elif str(cell_additional_value) == "4224":
                countKK += 1
                count42 += 1
                count24 += 1
                count4init += 1
                count4rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value) == "4334":
                countKK += 1
                count43 += 1
                count34 += 1
                count4init += 1
                count4rec += 1
                count3init += 1
                count3rec += 1
            elif str(cell_additional_value) == "4554":
                countKK += 1
                count45 += 1
                count54 += 1
                count4init += 1
                count4rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value) == "5115":
                countKK += 1
                count51 += 1
                count15 += 1
                count1init += 1
                count1rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value) == "5225":
                countKK += 1
                count52 += 1
                count25 += 1
                count5init += 1
                count5rec += 1
                count2init += 1
                count2rec += 1
            elif str(cell_additional_value) == "5335":
                countKK += 1
                count53 += 1
                count35 += 1
                count3init += 1
                count3rec += 1
                count5init += 1
                count5rec += 1
            elif str(cell_additional_value) == "5445":
                countKK += 1
                count54 += 1
                count45 += 1
                count5init += 1
                count5rec += 1
                count4init += 1
                count4rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'Z{start_row + i}'] = value

    workbook.save(excel_file)#
    print(f"Резултат подсчета лапки записан в сотлбец Z.")


    content_to_count = "nabeg     "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AA{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета набег записан в сотлбец AA.")


    content_to_count = "naprigivan"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AB{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета напрыгивание записан в сотлбец AB.")


    content_to_count = "obxvat    "

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AC{start_row + i}'] = value

    workbook.save(excel_file)#
    print(f"Резултат подсчета обхват записан в сотлбец AC.")


    content_to_count = "pogonya   "

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AD{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета погоня записан в сотлбец AD.")

     


    content_to_count = "priglashen"

    countMK = 0
    countM1 = 0
    countM2 = 0
    countM3 = 0
    countM4 = 0
    countM5 = 0

    count1init = 0
    count1rec= 0
    count2init = 0
    count2rec = 0
    count3init = 0
    count3rec = 0
    count4init = 0
    count4rec = 0
    count5init = 0
    count5rec = 0

    countKK = 0
    count12 = 0
    count13 = 0
    count14 = 0
    count15 = 0
    count21 = 0
    count23 = 0
    count24 = 0
    count25 = 0
    count31 = 0
    count32 = 0
    count34 = 0
    count35 = 0
    count41 = 0
    count42 = 0
    count43 = 0
    count45 = 0
    count51 = 0
    count52 = 0
    count53 = 0
    count54 = 0

    count1M = 0
    count2M = 0
    count3M = 0
    count4M = 0
    count5M = 0
    countKM = 0

    countTK = 0
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    countM = 0

    countN1 = 0
    countN2 = 0
    countN3 = 0
    countN4 = 0
    countN5 = 0
    countNK = 0
    countNM = 0

    count1N = 0
    count2N = 0
    count3N = 0
    count4N = 0
    count5N = 0
    countKN = 0
    countMN = 0

    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index]
        cell_additional_value = row[animal_column_index]
        cell_time = row[time_column_index]

        if cell_time is not None and float(cell_time) > end_time:
            break

        if str(cell_content) == content_to_count: 
            if str(cell_additional_value) == "M1":
                countMK += 1
                countM1 += 1
            elif str(cell_additional_value) == "M2":
                countMK += 1
                countM2 += 1
            elif str(cell_additional_value) == "M3":
                countMK += 1
                countM3 += 1
            elif str(cell_additional_value) == "M4":
                countMK += 1
                countM4 += 1
            elif str(cell_additional_value) == "M5":
                countMK += 1
                countM5 += 1
            elif str(cell_additional_value) == "Mkitten":
                countMK += 1
            elif str(cell_additional_value) == "MKitten":
                countMK += 1
            
            elif str(cell_additional_value) == "1M":
                countKM += 1
                count1M += 1
            elif str(cell_additional_value) == "2M":
                countKM += 1
                count2M += 1
            elif str(cell_additional_value) == "3M":
                countKM += 1
                count3M += 1
            elif str(cell_additional_value) == "4M":
                countKM += 1
                count4M += 1
            elif str(cell_additional_value) == "5M":
                countKM += 1
                count5M += 1
            elif str(cell_additional_value) == "kittenM":
                countKM += 1
            elif str(cell_additional_value) == "KittenM":
                countKM += 1
            
            elif str(cell_additional_value) == "1neighbour":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbour":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbour":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbour":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbour":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbour":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbour":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbour":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbour1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbour2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbour3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbour4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbour5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighbourM":
                countNM += 1
            elif str(cell_additional_value) == "neighbourKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighbourkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "1neighbor":
                count1N += 1
                countKN += 1
            elif str(cell_additional_value) == "2neighbor":
                countKN += 1
                count2N += 1
            elif str(cell_additional_value) == "3neighbor":
                countKN += 1
                count3N += 1
            elif str(cell_additional_value) == "4neighbor":
                countKN += 1
                count4N += 1
            elif str(cell_additional_value) == "5neighbor":
                countKN += 1
                count5N += 1
            elif str(cell_additional_value) == "Mneighbor":
                countMN += 1
            elif str(cell_additional_value) == "Kittenneighbor":
                countKN += 1
            elif str(cell_additional_value) == "kittenneighbor":
                countKN += 1
                
            elif str(cell_additional_value) == "neighbor1":
                countNK += 1
                countN1 += 1
            elif str(cell_additional_value) == "neighbor2":
                countNK += 1
                countN2 += 1
            elif str(cell_additional_value) == "neighbor3":
                countNK += 1
                countN3 += 1
            elif str(cell_additional_value) == "neighbor4":
                countNK += 1
                countN4 += 1
            elif str(cell_additional_value) == "neighbor5":
                countNK += 1
                countN5 += 1
            elif str(cell_additional_value) == "neighborM":
                countNM += 1
            elif str(cell_additional_value) == "neighborKitten":
                countNK += 1
            elif str(cell_additional_value) == "neighborkitten":
                countNK += 1
                            
            elif str(cell_additional_value) == "12":
                countKK += 1
                count12 += 1
                count1init += 1
                count2rec += 1
            elif str(cell_additional_value) == "13":
                countKK += 1
                count13 += 1
                count1init += 1
                count3rec += 1
            elif str(cell_additional_value) == "14":
                countKK += 1
                count14 += 1
                count1init += 1
                count4rec += 1
            elif str (cell_additional_value) == "15":
                countKK += 1
                count15 += 1
                count1init += 1
                count5rec += 1
            elif str(cell_additional_value) == "21":
                countKK += 1
                count21 += 1
                count2init += 1
                count1rec += 1
            elif str(cell_additional_value) == "23":
                countKK += 1
                count23 += 1
                count2init += 1
                count3rec += 1
            elif str(cell_additional_value) == "24":
                countKK += 1
                count24 += 1
                count2init += 1
                count4rec += 1
            elif str(cell_additional_value) == "25":
                countKK += 1
                count25 += 1
                count2init += 1
                count5rec += 1
            elif str(cell_additional_value) == "31":
                countKK += 1
                count31 += 1
                count3init += 1
                count1rec += 1
            elif str(cell_additional_value) == "32":
                countKK += 1
                count32 += 1
                count3init += 1
                count2rec += 1
            elif str(cell_additional_value) == "34":
                countKK += 1
                count34 += 1
                count3init += 1
                count4rec += 1
            elif str(cell_additional_value) == "35":
                countKK += 1
                count35 += 1
                count3init += 1
                count5rec += 1
            elif str(cell_additional_value) == "41":
                countKK += 1
                count41 += 1
                count4init += 1
                count1rec += 1
            elif str(cell_additional_value) == "42":
                countKK += 1
                count42 += 1
                count4init += 1
                count2rec += 1
            elif str(cell_additional_value) == "43":
                countKK += 1
                count43 += 1
                count4init += 1
                count3rec += 1
            elif str(cell_additional_value) == "45":
                countKK += 1
                count45 += 1
                count4init += 1
                count5rec += 1
            elif str(cell_additional_value) == "51":
                countKK += 1
                count51 += 1
                count5init += 1
                count1rec += 1
            elif str(cell_additional_value) == "52":
                countKK += 1
                count52 += 1
                count5init += 1
                count2rec += 1
            elif str(cell_additional_value) == "53":
                countKK += 1
                count53 += 1
                count5init += 1
                count3rec += 1
            elif str(cell_additional_value) == "54":
                countKK += 1
                count54 += 1
                count5init += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "1":
                countTK += 1
                count1 += 1
            elif str(cell_additional_value) == "2":
                countTK += 1
                count2 += 1
            elif str(cell_additional_value) == "3":
                countTK += 1
                count3 += 1
            elif str(cell_additional_value) == "4":
                countTK += 1
                count4 += 1
            elif str(cell_additional_value) == "5":
                countTK += 1
                count5 += 1
            elif str(cell_additional_value) == "M":
                countM += 1
            elif str(cell_additional_value) == "kitten":
                countTK += 1
            elif str(cell_additional_value) == "Kitten":
                countTK += 1
                
            elif str(cell_additional_value) == "1kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "1Kitten":
                countKK += 1
                count1init += 1
            elif str(cell_additional_value) == "kitten1":
                countKK += 1
                count1rec += 1
            elif str(cell_additional_value) == "Kitten1":
                countKK += 1
                count1rec += 1

            elif str(cell_additional_value) == "2kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "2Kitten":
                countKK += 1
                count2init += 1
            elif str(cell_additional_value) == "kitten2":
                countKK += 1
                count2rec += 1
            elif str(cell_additional_value) == "Kitten2":
                countKK += 1
                count2rec += 1
                
            elif str(cell_additional_value) == "3kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "3Kitten":
                countKK += 1
                count3init += 1
            elif str(cell_additional_value) == "kitten3":
                countKK += 1
                count3rec += 1
            elif str(cell_additional_value) == "Kitten3":
                countKK += 1
                count3rec += 1
                
            elif str(cell_additional_value) == "4kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "4Kitten":
                countKK += 1
                count4init += 1
            elif str(cell_additional_value) == "kitten4":
                countKK += 1
                count4rec += 1
            elif str(cell_additional_value) == "Kitten4":
                countKK += 1
                count4rec += 1
                
            elif str(cell_additional_value) == "5kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "5Kitten":
                countKK += 1
                count5init += 1
            elif str(cell_additional_value) == "kitten5":
                countKK += 1
                count5rec += 1
            elif str(cell_additional_value) == "Kitten5":
                countKK += 1
                count5rec += 1
                
    values_to_write = [countMK, countKM, countKK, countTK, countTK, countM, count1, count2, count3, count4, count5, count1init, count1rec, count2init, count2rec, count3init, count3rec, count4init, count4rec, count5init, count5rec, count12, count13, count14, count15, count21, count23, count24, count25, count31, count32, count34, count35, count41, count42, count43, count45, count51, count52, count53, count54, count1M, count2M, count3M, count4M, count5M, countM1, countM2, countM3, countM4, countM5, count1N, count2N, count3N, count4N, count5N, countKN, countN1, countN2, countN3, countN4, countN5, countNK, countMN, countNM]

    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AE{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Резултат подсчета приглашение записан в сотлбец AE.")


    def analyze_animal_play(excel_file):
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
        except FileNotFoundError:
            print(f"Ошибка: Файл {excel_file} не найден.")
            return
        except Exception as e:
            print(f"Ошибка при открытии файла: {e}")
            return

        games = []
        current_borba = None
        start_row = 8

        animal_pairs_duration = {}

        for index, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row - 1):
            try:
                action = str(row[3].value).strip().lower() if row[3].value is not None else ''
                animals = str(row[10].value).strip().lower() if row[10].value is not None else ''
                time = row[5].value

                if time is not None:
                    try:
                        time_float = float(time)
                        if time_float > end_time:
                            print(f"Время в строке {index + 1} превышает end_time ({end_time}). Цикл остановлен.")
                            break
                    except ValueError:
                        print(f"Ошибка: Невозможно преобразовать время в строке {index + 1} в число: {time}")
                        continue

                if 'borba' in action and not animals.endswith('end'):
                    if current_borba:
                        print(f"Борьба {current_borba['animals']} начата в строке {current_borba['start_row']}, но не завершена.")
                    current_borba = {
                        'start_time': time,
                        'animals': animals,
                        'start_row': index + 1
                    }

                elif 'borba' in action and animals.endswith('end'):
                    if current_borba:
                        if current_borba['animals'] == animals[:-3]:
                            current_borba['end_time'] = time
                            current_borba['end_row'] = index + 1

                            try:
                                start_time_float = float(current_borba['start_time'])
                                end_time_float = float(current_borba['end_time'])
                                current_borba['duration'] = round(end_time_float - start_time_float, 2)
                                pair_key = current_borba['animals']
                                if pair_key in animal_pairs_duration:
                                    animal_pairs_duration[pair_key] += current_borba['duration']
                                else:
                                    animal_pairs_duration[pair_key] = current_borba['duration']

                            except (ValueError, TypeError) as e:
                                print(f"Ошибка преобразования времени в строке {index + 1}: {e}")
                                current_borba['duration'] = None

                            games.append(current_borba)
                            current_borba = None

                        else:
                            print(f"Несоответствие участников борьбы. Ожидалось {current_borba['animals']}, получено {animals}.")
                    else:
                        print(f"Окончание борьбы без начала в строке {index + 1}.")

            except Exception as e:
                print(f"Ошибка в строке {index + 1}: {e}, Значение animals: {row[10].value}, Тип данных animals: {type(row[10].value)}")

        print(f"\nКоличество борьбы: {len(games)}")
        for i, borba in enumerate(games, 1):
            print(f"\nБорьба {i}:")
            print(f"Участники: {borba['animals']}")
            print(f"Начало борьбы: строка {borba['start_row']}, время {borba['start_time']} сек")
            print(f"Окончание борьбы: строка {borba['end_row']}, время {borba['end_time']} сек")
            print(f"Продолжительность борьбы: {borba['duration']} сек")

        output_row = 6
        for pair, duration in animal_pairs_duration.items():
            sheet[f'AG{output_row}'] = pair
            sheet[f'AH{output_row}'] = duration
            output_row += 1

        try:
            workbook.save(excel_file)
            print(f"\nДанные по группам записаны в столбцы AG и AH.")
        except Exception as e:
            print(f"Ошибка при сохранении файла: {e}")

    analyze_animal_play(excel_file)



    def analyze_animal_groups(excel_file):
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
        except FileNotFoundError:
            print(f"Ошибка: Файл {excel_file} не найден.")
            return
        except Exception as e:
            print(f"Ошибка при открытии файла: {e}")
            return

        total_duration = 0
        group1_duration = 0
        group2_duration = 0
        group3_duration = 0

        start_row = 6

        if sheet.max_column < 34:
            print("Ошибка: В файле недостаточно столбцов.")
            return

        for row in sheet.iter_rows(min_row=start_row, min_col=33, max_col=34, values_only=True):
            animals = str(row[0]) if row[0] is not None else ''
            duration = row[1]

            if isinstance(duration, (int, float)):
                total_duration += duration

                if animals.isdigit():
                    group1_duration += duration
                elif animals.lower().startswith('m'):
                    group2_duration += duration
                elif animals.lower().endswith('m'):
                    group3_duration += duration
                    print(f"+KimM")

        sheet['AJ6'] = total_duration
        sheet['AJ7'] = group1_duration
        sheet['AJ8'] = group2_duration
        sheet['AJ9'] = group3_duration

        try:
            workbook.save(excel_file)
            print(f"Результаты успешно записаны в столбец AJ.")
        except Exception as e:
            print(f"Ошибка при сохранении файла: {e}")

    analyze_animal_groups(excel_file)

    
    column_letter = 'D'

    content_to_count = "allogrumin"
    output_cell = "P10"

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    print(f"Активный лист: {sheet.title}")

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            print(f"Время в строке {row[0].row} превышает end_time ({end_time}). Остановка обработки.")
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "gruming   "
    output_cell = "Q10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "igra      "
    output_cell = "R10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "spredmetom"
    output_cell = "V10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "bokom     "
    output_cell = "W10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "ckradivan "
    output_cell = "X10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "zataivanie"
    output_cell = "Y10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "lapki     "
    output_cell = "Z10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "nabeg     "
    output_cell = "AA10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "naprigivan"
    output_cell = "AB10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "obxvat    "
    output_cell = "AC10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "pogonya   "
    output_cell = "AD10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    content_to_count = "priglashen"
    output_cell = "AE10"

    count = 0
    stop_processing = False

    for row in sheet.iter_rows(min_row=7):
        time_cell = row[5].value
        if time_cell is not None and float(time_cell) > end_time:
            stop_processing = True
            break

        cell = row[openpyxl.utils.column_index_from_string(column_letter) - 1]
        if cell.value == content_to_count:
            count += 1

    sheet[output_cell] = count
    workbook.save(excel_file)
    print(f"Общее количество '{content_to_count}': {count}, результат записан в ячейку {output_cell}.")


    ws = workbook.active

    values = [ws['S11'].value, ws['S12'].value, ws['S13'].value, 
              ws['S14'].value, ws['S15'].value, ws['S16'].value]
    total = sum(v for v in values if v is not None)

    ws['S10'] = total
    workbook.save(excel_file)

    print(f"Общее количество selfplay: {total}, резултат записан в ячейку S10.")


    values = [ws['T6'].value, ws['T7'].value]
    total = sum(v for v in values if v is not None)

    ws['T10'] = total
    workbook.save(excel_file)

    print(f"Общее количество игра с активной мамой: {total}, резултат записан в ячейку T10.")



    values = [ws['U6'].value, ws['U7'].value]
    total = sum(v for v in values if v is not None)

    ws['U10'] = total
    workbook.save(excel_file)

    print(f"Общее количество игра с пассивной мамой: {total}, резултат записан в ячейку U10.")



    values = [ws['T10'].value, ws['U10'].value]
    total = sum(v for v in values if v is not None)

    ws['U4'] = total
    workbook.save(excel_file)


    variable = "igrasmamoj"

    ws['U3'] = variable
    workbook.save(excel_file)

    print(f"Общее количество игра c мамой (акт + пас): {total}, резултат записан в ячейку U4.")
    
    
#     NEVIDNO
#         1 KITTEN

    if kit_number == 1:
        print("выводок из 1 котенка")
        
        output_cell = "S2"
     
        workbook = load_workbook(excel_file, data_only=True)
        sheet = workbook.active
       
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        total_time = 0
        kit_sitting_start_time = 0.0
        is_invisible = True
        first_event_processed = False

        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "1neighbor", "neighbour1", "neighbor1", "1neighbourneighbour1",
            "1neighborneighbor1", "neighbour11neighbour", "neighbor11neighbor",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }

        non_interrupting_actions = {"nevidno", "vokal"}
        last_valid_time = None

        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            try:
                action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
                animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
                time = row[openpyxl.utils.column_index_from_string(time_column) - 1]
            
                if not action or not animal or time is None:
                    continue

                current_time = float(time)
                last_valid_time = current_time  
            
                if current_time > end_time:
                    break
                
                is_kit_event = (animal == "1" or animal in kit1_aliases)
                
                if is_invisible:
                    if is_kit_event and action not in non_interrupting_actions:
                        period_end = min(end_time, current_time)
                        total_time += period_end - kit_sitting_start_time
                        is_invisible = False
                        kit_sitting_start_time = None
                        first_event_processed = True
                    
                    elif is_kit_event and action == "nevidno":
                        pass
                        
                else:
                    if is_kit_event and action == "nevidno":
                        kit_sitting_start_time = current_time
                        is_invisible = True

            except (ValueError, TypeError):
                continue

        if is_invisible and kit_sitting_start_time is not None:
            if last_valid_time is not None:
                period_end = min(end_time, last_valid_time)
                total_time += period_end - kit_sitting_start_time
            else:
                total_time += end_time

        if last_valid_time is None and end_time > 0:
            total_time = end_time

        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время невидно 1 котенка (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")



#         2 KITTENS
    if kit_number == 2:
        print("выводок из 2 котят")
        
        output_cell = "S2"
        
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        workbook = load_workbook(excel_file)
        sheet = workbook.active
            
        total_time = 0
        kit1_nevidno = True
        kit2_nevidno = True
        simultaneous_nevidno_start_time = 0.0
        last_valid_time = None

        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
            "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
            "kitten1", "kitten1", "1kitten", "1kitten",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }
        
        kit2_aliases = {
            "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
            "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
            "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
            "kitten2", "kitten2", "2kitten", "2kitten",
            "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
            "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
            "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
            "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
            "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
        }
        
        both_kits_aliases = {
            "12", "21", "1221", "2112"
        }
        
        non_interrupting_actions = {"nevidno", "vokal"}

        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            try:
                action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
                animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
                time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

                if not action or not animal or time is None:
                    continue

                current_time = float(time)
                last_valid_time = current_time
                
                if current_time > end_time:
                    break

                is_kit1 = animal in kit1_aliases or animal == "1"
                is_kit2 = animal in kit2_aliases or animal == "2"
                is_both = animal in both_kits_aliases
                
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        if is_kit1:
                            kit1_nevidno = True
                        if is_kit2:
                            kit2_nevidno = True
                        if is_both:
                            kit1_nevidno = True
                            kit2_nevidno = True
                else:
                    if is_kit1:
                        kit1_nevidno = False
                    if is_kit2:
                        kit2_nevidno = False
                    if is_both:
                        kit1_nevidno = False
                        kit2_nevidno = False

                all_invisible = kit1_nevidno and kit2_nevidno
                
                if all_invisible:
                    if simultaneous_nevidno_start_time is None:
                        simultaneous_nevidno_start_time = current_time
                else:
                    if simultaneous_nevidno_start_time is not None:
                        period_end = min(end_time, current_time)
                        duration = period_end - simultaneous_nevidno_start_time
                        if duration > 0:
                            total_time += duration
                        simultaneous_nevidno_start_time = None

            except (ValueError, TypeError):
                continue

        if simultaneous_nevidno_start_time is not None:
            if last_valid_time is not None:
                period_end = min(end_time, last_valid_time)
                duration = period_end - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
            else:
                total_time += end_time

        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время одновременного невидно для 2 котят (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")



#         3 KITTENS
    
    if kit_number == 3:
        print("выводок из 3 котят")
        
        output_cell = "S2"
        
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        workbook = load_workbook(excel_file)
        sheet = workbook.active
            
        total_time = 0
        kit1_nevidno = True
        kit2_nevidno = True
        kit3_nevidno = True
        simultaneous_nevidno_start_time = 0.0
        last_valid_time = None

        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
            "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
            "kitten1", "1kitten",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }
        
        kit2_aliases = {
            "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
            "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
            "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
            "kitten2", "2kitten",
            "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
            "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
            "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
            "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
            "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
        }
        
        kit3_aliases = {
            "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
            "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
            "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
            "kitten3", "3kitten",
            "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
            "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
            "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
            "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
            "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
        }
        
        kit12_aliases = {"12", "21", "1221", "2112"}
        kit13_aliases = {"13", "31", "1331", "3113"}
        kit23_aliases = {"23", "32", "2332", "3223"}
        kit123_aliases = {"123", "132", "213", "231", "312", "321"}
        
        non_interrupting_actions = {"nevidno", "vokal"}

        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            try:
                action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
                animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
                time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

                if not action or not animal or time is None:
                    continue

                current_time = float(time)
                last_valid_time = current_time
                
                if current_time > end_time:
                    break

                is_kit1 = animal in kit1_aliases or animal == "1"
                is_kit2 = animal in kit2_aliases or animal == "2"
                is_kit3 = animal in kit3_aliases or animal == "3"
                is_kit12 = animal in kit12_aliases
                is_kit13 = animal in kit13_aliases
                is_kit23 = animal in kit23_aliases
                is_kit123 = animal in kit123_aliases
                
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        if is_kit1:
                            kit1_nevidno = True
                        if is_kit2:
                            kit2_nevidno = True
                        if is_kit3:
                            kit3_nevidno = True
                        if is_kit12:
                            kit1_nevidno = True
                            kit2_nevidno = True
                        if is_kit13:
                            kit1_nevidno = True
                            kit3_nevidno = True
                        if is_kit23:
                            kit2_nevidno = True
                            kit3_nevidno = True
                        if is_kit123:
                            kit1_nevidno = True
                            kit2_nevidno = True
                            kit3_nevidno = True
                else:
                    if is_kit1:
                        kit1_nevidno = False
                    if is_kit2:
                        kit2_nevidno = False
                    if is_kit3:
                        kit3_nevidno = False
                    if is_kit12:
                        kit1_nevidno = False
                        kit2_nevidno = False
                    if is_kit13:
                        kit1_nevidno = False
                        kit3_nevidno = False
                    if is_kit23:
                        kit2_nevidno = False
                        kit3_nevidno = False
                    if is_kit123:
                        kit1_nevidno = False
                        kit2_nevidno = False
                        kit3_nevidno = False

                all_invisible = kit1_nevidno and kit2_nevidno and kit3_nevidno
                
                if all_invisible:
                    if simultaneous_nevidno_start_time is None:
                        simultaneous_nevidno_start_time = current_time
                else:
                    if simultaneous_nevidno_start_time is not None:
                        period_end = min(end_time, current_time)
                        duration = period_end - simultaneous_nevidno_start_time
                        if duration > 0:
                            total_time += duration
                        simultaneous_nevidno_start_time = None

            except (ValueError, TypeError):
                continue

        if simultaneous_nevidno_start_time is not None:
            if last_valid_time is not None:
                period_end = min(end_time, last_valid_time)
                duration = period_end - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
            else:
                total_time += end_time

        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время одновременного невидно для 3 котят (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")



#         4 KITTENS

    if kit_number == 4:
        print("выводок из 4 котят")
        
        output_cell = "S2"
        
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        workbook = load_workbook(excel_file)
        sheet = workbook.active
            
        total_time = 0
        kit1_nevidno = True
        kit2_nevidno = True
        kit3_nevidno = True
        kit4_nevidno = True
        simultaneous_nevidno_start_time = 0.0
        last_valid_time = None

        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
            "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
            "kitten1", "1kitten",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }
        
        kit2_aliases = {
            "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
            "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
            "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
            "kitten2", "2kitten",
            "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
            "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
            "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
            "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
            "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
        }
        
        kit3_aliases = {
            "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
            "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
            "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
            "kitten3", "3kitten",
            "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
            "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
            "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
            "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
            "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
        }
        
        kit4_aliases = {
            "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
            "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
            "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
            "kitten4", "4kitten",
            "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
            "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
            "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
            "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
            "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
        }
        
        kit12_aliases = {"12", "21", "1221", "2112"}
        kit13_aliases = {"13", "31", "1331", "3113"}
        kit14_aliases = {"14", "41", "1441", "4114"}
        kit23_aliases = {"23", "32", "2332", "3223"}
        kit24_aliases = {"24", "42", "2442", "4224"}
        kit34_aliases = {"34", "43", "3443", "4334"}
        
        kit123_aliases = {"123", "132", "213", "231", "312", "321"}
        kit124_aliases = {"124", "142", "214", "241", "412", "421"}
        kit134_aliases = {"134", "143", "314", "341", "413", "431"}
        kit234_aliases = {"234", "243", "324", "342", "423", "432"}
        
        kit1234_aliases = {"1234", "1243", "1324", "1342", "1423", "1432", 
                           "2134", "2143", "2314", "2341", "2413", "2431",
                           "3124", "3142", "3214", "3241", "3412", "3421", 
                           "4123", "4132", "4213", "4231", "4312", "4321"}
        
        non_interrupting_actions = {"nevidno", "vokal"}

        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            try:
                action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
                animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
                time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

                if not action or not animal or time is None:
                    continue

                current_time = float(time)
                last_valid_time = current_time
                
                if current_time > end_time:
                    break

                is_kit1 = animal in kit1_aliases or animal == "1"
                is_kit2 = animal in kit2_aliases or animal == "2"
                is_kit3 = animal in kit3_aliases or animal == "3"
                is_kit4 = animal in kit4_aliases or animal == "4"
                
                is_kit12 = animal in kit12_aliases
                is_kit13 = animal in kit13_aliases
                is_kit14 = animal in kit14_aliases
                is_kit23 = animal in kit23_aliases
                is_kit24 = animal in kit24_aliases
                is_kit34 = animal in kit34_aliases
                
                is_kit123 = animal in kit123_aliases
                is_kit124 = animal in kit124_aliases
                is_kit134 = animal in kit134_aliases
                is_kit234 = animal in kit234_aliases
                
                is_kit1234 = animal in kit1234_aliases
                
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        if is_kit1:
                            kit1_nevidno = True
                        if is_kit2:
                            kit2_nevidno = True
                        if is_kit3:
                            kit3_nevidno = True
                        if is_kit4:
                            kit4_nevidno = True
                            
                        if is_kit12:
                            kit1_nevidno = True
                            kit2_nevidno = True
                        if is_kit13:
                            kit1_nevidno = True
                            kit3_nevidno = True
                        if is_kit14:
                            kit1_nevidno = True
                            kit4_nevidno = True
                        if is_kit23:
                            kit2_nevidno = True
                            kit3_nevidno = True
                        if is_kit24:
                            kit2_nevidno = True
                            kit4_nevidno = True
                        if is_kit34:
                            kit3_nevidno = True
                            kit4_nevidno = True
                            
                        if is_kit123:
                            kit1_nevidno = True
                            kit2_nevidno = True
                            kit3_nevidno = True
                        if is_kit124:
                            kit1_nevidno = True
                            kit2_nevidno = True
                            kit4_nevidno = True
                        if is_kit134:
                            kit1_nevidno = True
                            kit3_nevidno = True
                            kit4_nevidno = True
                        if is_kit234:
                            kit2_nevidno = True
                            kit3_nevidno = True
                            kit4_nevidno = True
                            
                        if is_kit1234:
                            kit1_nevidno = True
                            kit2_nevidno = True
                            kit3_nevidno = True
                            kit4_nevidno = True
                else:
                    if is_kit1:
                        kit1_nevidno = False
                    if is_kit2:
                        kit2_nevidno = False
                    if is_kit3:
                        kit3_nevidno = False
                    if is_kit4:
                        kit4_nevidno = False
                        
                    if is_kit12:
                        kit1_nevidno = False
                        kit2_nevidno = False
                    if is_kit13:
                        kit1_nevidno = False
                        kit3_nevidno = False
                    if is_kit14:
                        kit1_nevidno = False
                        kit4_nevidno = False
                    if is_kit23:
                        kit2_nevidno = False
                        kit3_nevidno = False
                    if is_kit24:
                        kit2_nevidno = False
                        kit4_nevidno = False
                    if is_kit34:
                        kit3_nevidno = False
                        kit4_nevidno = False
                        
                    if is_kit123:
                        kit1_nevidno = False
                        kit2_nevidno = False
                        kit3_nevidno = False
                    if is_kit124:
                        kit1_nevidno = False
                        kit2_nevidno = False
                        kit4_nevidno = False
                    if is_kit134:
                        kit1_nevidno = False
                        kit3_nevidno = False
                        kit4_nevidno = False
                    if is_kit234:
                        kit2_nevidno = False
                        kit3_nevidno = False
                        kit4_nevidno = False
                        
                    if is_kit1234:
                        kit1_nevidno = False
                        kit2_nevidno = False
                        kit3_nevidno = False
                        kit4_nevidno = False

                all_invisible = kit1_nevidno and kit2_nevidno and kit3_nevidno and kit4_nevidno
                
                if all_invisible:
                    if simultaneous_nevidno_start_time is None:
                        simultaneous_nevidno_start_time = current_time
                else:
                    if simultaneous_nevidno_start_time is not None:
                        period_end = min(end_time, current_time)
                        duration = period_end - simultaneous_nevidno_start_time
                        if duration > 0:
                            total_time += duration
                        simultaneous_nevidno_start_time = None

            except (ValueError, TypeError):
                continue

        if simultaneous_nevidno_start_time is not None:
            if last_valid_time is not None:
                period_end = min(end_time, last_valid_time)
                duration = period_end - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
            else:
                total_time += end_time

        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время одновременного невидно для 4 котят (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")



#         5 KITTENS

    if kit_number == 5:
        print("выводок из 5 котят")
        
        output_cell = "S2"
        
        action_column = 'D'
        animal_column = 'K'
        time_column = 'F'

        workbook = load_workbook(excel_file)
        sheet = workbook.active
            
        total_time = 0
        kit1_nevidno = True
        kit2_nevidno = True
        kit3_nevidno = True
        kit4_nevidno = True
        kit5_nevidno = True
        simultaneous_nevidno_start_time = 0.0
        last_valid_time = None

        kit1_aliases = {
            "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
            "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
            "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
            "kitten1", "1kitten",
            "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
            "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
            "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
            "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
            "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
        }
        
        kit2_aliases = {
            "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
            "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
            "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
            "kitten2", "2kitten",
            "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
            "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
            "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
            "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
            "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
        }
        
        kit3_aliases = {
            "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
            "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
            "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
            "kitten3", "3kitten",
            "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
            "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
            "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
            "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
            "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
        }
        
        kit4_aliases = {
            "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
            "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
            "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
            "kitten4", "4kitten",
            "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
            "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
            "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
            "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
            "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
        }
        
        kit5_aliases = {
            "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
            "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
            "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
            "kitten5", "5kitten",
            "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
            "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
            "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
            "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
            "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
        }
        
        kit12_aliases = {"12", "21", "1221", "2112"}
        kit13_aliases = {"13", "31", "1331", "3113"}
        kit14_aliases = {"14", "41", "1441", "4114"}
        kit15_aliases = {"15", "51", "1551", "5115"}
        kit23_aliases = {"23", "32", "2332", "3223"}
        kit24_aliases = {"24", "42", "2442", "4224"}
        kit25_aliases = {"25", "52", "2552", "5225"}
        kit34_aliases = {"34", "43", "3443", "4334"}
        kit35_aliases = {"35", "53", "3553", "5335"}
        kit45_aliases = {"45", "54", "4554", "5445"}
        
        kit123_aliases = {"123", "132", "213", "231", "312", "321"}
        kit124_aliases = {"124", "142", "214", "241", "412", "421"}
        kit125_aliases = {"125", "152", "215", "251", "512", "521"}
        kit134_aliases = {"134", "143", "314", "341", "413", "431"}
        kit135_aliases = {"135", "153", "315", "351", "513", "531"}
        kit145_aliases = {"145", "154", "415", "451", "514", "541"}
        kit234_aliases = {"234", "243", "324", "342", "423", "432"}
        kit235_aliases = {"235", "253", "325", "352", "523", "532"}
        kit245_aliases = {"245", "254", "425", "452", "524", "542"}
        kit345_aliases = {"345", "354", "435", "453", "534", "543"}
        
        kit1234_aliases = {"1234", "1243", "1324", "1342", "1423", "1432",
                           "2134", "2143", "2314", "2341", "2413", "2431",
                           "3124", "3142", "3214", "3241", "3412", "3421",
                           "4123", "4132", "4213", "4231", "4312", "4321"}
        
        kit1235_aliases = {"1235", "1253", "1325", "1352", "1523", "1532",
                           "2135", "2153", "2315", "2351", "2513", "2531",
                           "3125", "3152", "3215", "3251", "3512", "3521",
                           "5123", "5132", "5213", "5231", "5312", "5321"}
        
        kit1245_aliases = {"1245", "1254", "1425", "1452", "1524", "1542",
                           "2145", "2154", "2415", "2451", "2514", "2541",
                           "4125", "4152", "4215", "4251", "4512", "4521",
                           "5124", "5142", "5214", "5241", "5412", "5421"}
        
        kit1345_aliases = {"1345", "1354", "1435", "1453", "1534", "1543",
                           "3145", "3154", "3415", "3451", "3514", "3541",
                           "4135", "4153", "4315", "4351", "4513", "4531",
                           "5134", "5143", "5314", "5341", "5413", "5431"}
        
        kit2345_aliases = {"2345", "2354", "2435", "2453", "2534", "2543",
                           "3245", "3254", "3425", "3452", "3524", "3542",
                           "4235", "4253", "4325", "4352", "4523", "4532",
                           "5234", "5243", "5324", "5342", "5423", "5432"}
        
        kit12345_aliases = {"12345", "12354", "12435", "12453", "12534", "12543",
                            "13245", "13254", "13425", "13452", "13524", "13542",
                            "14235", "14253", "14325", "14352", "14523", "14532",
                            "15234", "15243", "15324", "15342", "15423", "15432",
                            "21345", "21354", "21435", "21453", "21534", "21543",
                            "23145", "23154", "23415", "23451", "23514", "23541",
                            "24135", "24153", "24315", "24351", "24513", "24531",
                            "25134", "25143", "25314", "25341", "25413", "25431",
                            "31245", "31254", "31425", "31452", "31524", "31542",
                            "32145", "32154", "32415", "32451", "32514", "32541",
                            "34125", "34152", "34215", "34251", "34512", "34521",
                            "35124", "35142", "35214", "35241", "35412", "35421",
                            "41235", "41253", "41325", "41352", "41523", "41532",
                            "42135", "42153", "42315", "42351", "42513", "42531",
                            "43125", "43152", "43215", "43251", "43512", "43521",
                            "45123", "45132", "45213", "45231", "45312", "45321",
                            "51234", "51243", "51324", "51342", "51423", "51432",
                            "52134", "52143", "52314", "52341", "52413", "52431",
                            "53124", "53142", "53214", "53241", "53412", "53421",
                            "54123", "54132", "54213", "54231", "54312", "54321"}
        
        non_interrupting_actions = {"nevidno", "vokal"}

        for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            try:
                action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
                animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
                time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

                if not action or not animal or time is None:
                    continue

                current_time = float(time)
                last_valid_time = current_time
                
                if current_time > end_time:
                    break

                is_kit1 = animal in kit1_aliases or animal == "1"
                is_kit2 = animal in kit2_aliases or animal == "2"
                is_kit3 = animal in kit3_aliases or animal == "3"
                is_kit4 = animal in kit4_aliases or animal == "4"
                is_kit5 = animal in kit5_aliases or animal == "5"
                
                is_kit12 = animal in kit12_aliases
                is_kit13 = animal in kit13_aliases
                is_kit14 = animal in kit14_aliases
                is_kit15 = animal in kit15_aliases
                is_kit23 = animal in kit23_aliases
                is_kit24 = animal in kit24_aliases
                is_kit25 = animal in kit25_aliases
                is_kit34 = animal in kit34_aliases
                is_kit35 = animal in kit35_aliases
                is_kit45 = animal in kit45_aliases
                
                is_kit123 = animal in kit123_aliases
                is_kit124 = animal in kit124_aliases
                is_kit125 = animal in kit125_aliases
                is_kit134 = animal in kit134_aliases
                is_kit135 = animal in kit135_aliases
                is_kit145 = animal in kit145_aliases
                is_kit234 = animal in kit234_aliases
                is_kit235 = animal in kit235_aliases
                is_kit245 = animal in kit245_aliases
                is_kit345 = animal in kit345_aliases
                
                is_kit1234 = animal in kit1234_aliases
                is_kit1235 = animal in kit1235_aliases
                is_kit1245 = animal in kit1245_aliases
                is_kit1345 = animal in kit1345_aliases
                is_kit2345 = animal in kit2345_aliases
                
                is_kit12345 = animal in kit12345_aliases
                
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        if is_kit1:
                            kit1_nevidno = True
                        if is_kit2:
                            kit2_nevidno = True
                        if is_kit3:
                            kit3_nevidno = True
                        if is_kit4:
                            kit4_nevidno = True
                        if is_kit5:
                            kit5_nevidno = True
                        
                        if is_kit12:
                            kit1_nevidno = True; kit2_nevidno = True
                        if is_kit13:
                            kit1_nevidno = True; kit3_nevidno = True
                        if is_kit14:
                            kit1_nevidno = True; kit4_nevidno = True
                        if is_kit15:
                            kit1_nevidno = True; kit5_nevidno = True
                        if is_kit23:
                            kit2_nevidno = True; kit3_nevidno = True
                        if is_kit24:
                            kit2_nevidno = True; kit4_nevidno = True
                        if is_kit25:
                            kit2_nevidno = True; kit5_nevidno = True
                        if is_kit34:
                            kit3_nevidno = True; kit4_nevidno = True
                        if is_kit35:
                            kit3_nevidno = True; kit5_nevidno = True
                        if is_kit45:
                            kit4_nevidno = True; kit5_nevidno = True
                        
                        if is_kit123:
                            kit1_nevidno = True; kit2_nevidno = True; kit3_nevidno = True
                        if is_kit124:
                            kit1_nevidno = True; kit2_nevidno = True; kit4_nevidno = True
                        if is_kit125:
                            kit1_nevidno = True; kit2_nevidno = True; kit5_nevidno = True
                        if is_kit134:
                            kit1_nevidno = True; kit3_nevidno = True; kit4_nevidno = True
                        if is_kit135:
                            kit1_nevidno = True; kit3_nevidno = True; kit5_nevidno = True
                        if is_kit145:
                            kit1_nevidno = True; kit4_nevidno = True; kit5_nevidno = True
                        if is_kit234:
                            kit2_nevidno = True; kit3_nevidno = True; kit4_nevidno = True
                        if is_kit235:
                            kit2_nevidno = True; kit3_nevidno = True; kit5_nevidno = True
                        if is_kit245:
                            kit2_nevidno = True; kit4_nevidno = True; kit5_nevidno = True
                        if is_kit345:
                            kit3_nevidno = True; kit4_nevidno = True; kit5_nevidno = True
                        
                        if is_kit1234:
                            kit1_nevidno = True; kit2_nevidno = True; kit3_nevidno = True; kit4_nevidno = True
                        if is_kit1235:
                            kit1_nevidno = True; kit2_nevidno = True; kit3_nevidno = True; kit5_nevidno = True
                        if is_kit1245:
                            kit1_nevidno = True; kit2_nevidno = True; kit4_nevidno = True; kit5_nevidno = True
                        if is_kit1345:
                            kit1_nevidno = True; kit3_nevidno = True; kit4_nevidno = True; kit5_nevidno = True
                        if is_kit2345:
                            kit2_nevidno = True; kit3_nevidno = True; kit4_nevidno = True; kit5_nevidno = True
                        
                        if is_kit12345:
                            kit1_nevidno = True; kit2_nevidno = True; kit3_nevidno = True; kit4_nevidno = True; kit5_nevidno = True
                else:
                    if is_kit1:
                        kit1_nevidno = False
                    if is_kit2:
                        kit2_nevidno = False
                    if is_kit3:
                        kit3_nevidno = False
                    if is_kit4:
                        kit4_nevidno = False
                    if is_kit5:
                        kit5_nevidno = False
                    
                    if is_kit12:
                        kit1_nevidno = False; kit2_nevidno = False
                    if is_kit13:
                        kit1_nevidno = False; kit3_nevidno = False
                    if is_kit14:
                        kit1_nevidno = False; kit4_nevidno = False
                    if is_kit15:
                        kit1_nevidno = False; kit5_nevidno = False
                    if is_kit23:
                        kit2_nevidno = False; kit3_nevidno = False
                    if is_kit24:
                        kit2_nevidno = False; kit4_nevidno = False
                    if is_kit25:
                        kit2_nevidno = False; kit5_nevidno = False
                    if is_kit34:
                        kit3_nevidno = False; kit4_nevidno = False
                    if is_kit35:
                        kit3_nevidno = False; kit5_nevidno = False
                    if is_kit45:
                        kit4_nevidno = False; kit5_nevidno = False
                    
                    if is_kit123:
                        kit1_nevidno = False; kit2_nevidno = False; kit3_nevidno = False
                    if is_kit124:
                        kit1_nevidno = False; kit2_nevidno = False; kit4_nevidno = False
                    if is_kit125:
                        kit1_nevidno = False; kit2_nevidno = False; kit5_nevidno = False
                    if is_kit134:
                        kit1_nevidno = False; kit3_nevidno = False; kit4_nevidno = False
                    if is_kit135:
                        kit1_nevidno = False; kit3_nevidno = False; kit5_nevidno = False
                    if is_kit145:
                        kit1_nevidno = False; kit4_nevidno = False; kit5_nevidno = False
                    if is_kit234:
                        kit2_nevidno = False; kit3_nevidno = False; kit4_nevidno = False
                    if is_kit235:
                        kit2_nevidno = False; kit3_nevidno = False; kit5_nevidno = False
                    if is_kit245:
                        kit2_nevidno = False; kit4_nevidno = False; kit5_nevidno = False
                    if is_kit345:
                        kit3_nevidno = False; kit4_nevidno = False; kit5_nevidno = False
                    
                    if is_kit1234:
                        kit1_nevidno = False; kit2_nevidno = False; kit3_nevidno = False; kit4_nevidno = False
                    if is_kit1235:
                        kit1_nevidno = False; kit2_nevidno = False; kit3_nevidno = False; kit5_nevidno = False
                    if is_kit1245:
                        kit1_nevidno = False; kit2_nevidno = False; kit4_nevidno = False; kit5_nevidno = False
                    if is_kit1345:
                        kit1_nevidno = False; kit3_nevidno = False; kit4_nevidno = False; kit5_nevidno = False
                    if is_kit2345:
                        kit2_nevidno = False; kit3_nevidno = False; kit4_nevidno = False; kit5_nevidno = False
                    
                    if is_kit12345:
                        kit1_nevidno = False; kit2_nevidno = False; kit3_nevidno = False; kit4_nevidno = False; kit5_nevidno = False

                all_invisible = kit1_nevidno and kit2_nevidno and kit3_nevidno and kit4_nevidno and kit5_nevidno
                
                if all_invisible:
                    if simultaneous_nevidno_start_time is None:
                        simultaneous_nevidno_start_time = current_time
                else:
                    if simultaneous_nevidno_start_time is not None:
                        period_end = min(end_time, current_time)
                        duration = period_end - simultaneous_nevidno_start_time
                        if duration > 0:
                            total_time += duration
                        simultaneous_nevidno_start_time = None

            except (ValueError, TypeError):
                continue

        if simultaneous_nevidno_start_time is not None:
            if last_valid_time is not None:
                period_end = min(end_time, last_valid_time)
                duration = period_end - simultaneous_nevidno_start_time
                if duration > 0:
                    total_time += duration
            else:
                total_time += end_time

        sheet[output_cell] = total_time
        workbook.save(excel_file)
        print(f"Общее время одновременного невидно для 5 котят (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейку {output_cell}.")



    #TotalTime
    output_cell = "R2"

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    sheet[output_cell] = end_time
    workbook.save(excel_file)

    print(f"Значение TotlTime записано в ячейку {output_cell}.")
        
        
    #observed
    cell_tt = "R2"
    cell_nv = "S2"
    output_cell = "T2"

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    value_tt = float(sheet[cell_tt].value)
    value_nv = float(sheet[cell_nv].value)

    observed = value_tt - value_nv

    sheet[output_cell] = observed
    workbook.save(excel_file)

    print(f"Значение observed: {observed}, записана в ячейку {output_cell}.")


    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    content_to_count = "cocanie   "

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    countMS = 0
    countKS = 0
    countST = 0
    countCS = 0
    countCUS = 0

    count1SS = 0
    count2SS = 0
    count3SS = 0
    count4SS = 0
    countNSS = 0

    count1USS = 0
    count2USS = 0
    count3USS = 0
    count4USS = 0
    countNUSS = 0


    action_column_index = openpyxl.utils.column_index_from_string(action_column) - 1
    animal_column_index = openpyxl.utils.column_index_from_string(animal_column) - 1
    time_column_index = openpyxl.utils.column_index_from_string(time_column) - 1

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index]
        cell_time_value = row[time_column_index]
        
        if cell_time_value is not None and float(cell_time_value) > end_time:
            break

        if cell_content == content_to_count: 
            if str(cell_additional_value) == "1 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value) == "1 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value) == "1 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value) == "1 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value) == "1 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
                
            elif str(cell_additional_value) == "2 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value) == "2 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value) == "2 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value) == "2 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value) == "2 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
                
            elif str(cell_additional_value) == "3 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value) == "3 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value) == "3 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value) == "3 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value) == "3 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
            
            elif str(cell_additional_value) == "4 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value) == "4 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value) == "4 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value) == "4 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value) == "4 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
                
            elif str(cell_additional_value) == "5 1pairtrysuccess":
                countCS += 1
                count1SS += 1
            elif str(cell_additional_value) == "5 2pairtrysuccess":
                countCS += 1
                count2SS += 1
            elif str(cell_additional_value) == "5 3pairtrysuccess":
                countCS += 1
                count3SS += 1
            elif str(cell_additional_value) == "5 4pairtrysuccess":
                countCS += 1
                count4SS += 1
            elif str(cell_additional_value) == "5 ?pairtrysuccess":
                countCS += 1
                countNSS += 1
                
            elif str(cell_additional_value) == "1 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value) == "1 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value) == "1 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value) == "1 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value) == "1 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
            
            elif str(cell_additional_value) == "2 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value) == "2 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value) == "2 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value) == "2 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value) == "2 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
                
            elif str(cell_additional_value) == "3 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value) == "3 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value) == "3 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value) == "3 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value) == "3 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
                
            elif str(cell_additional_value) == "4 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value) == "4 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value) == "4 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value) == "4 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value) == "4 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
                
            elif str(cell_additional_value) == "5 1pairtryunsuccess":
                countCUS += 1
                count1USS += 1
            elif str(cell_additional_value) == "5 2pairtryunsuccess":
                countCUS += 1
                count2USS += 1
            elif str(cell_additional_value) == "5 3pairtryunsuccess":
                countCUS += 1
                count3USS += 1
            elif str(cell_additional_value) == "5 4pairtryunsuccess":
                countCUS += 1
                count4USS += 1
            elif str(cell_additional_value) == "5 ?pairtryunsuccess":
                countCUS += 1
                countNUSS += 1
                
            elif str(cell_additional_value) == "1 1pairend":
                countST += 1
            elif str(cell_additional_value) == "1 2pairend":
                countST += 1
            elif str(cell_additional_value) == "1 3pairend":
                countST += 1
            elif str(cell_additional_value) == "1 4pairend":
                countST += 1
            elif str(cell_additional_value) == "1 ?pairend":
                countST += 1
                
            elif str(cell_additional_value) == "2 1pairend":
                countST += 1
            elif str(cell_additional_value) == "2 2pairend":
                countST += 1
            elif str(cell_additional_value) == "2 3pairend":
                countST += 1
            elif str(cell_additional_value) == "2 4pairend":
                countST += 1
            elif str(cell_additional_value) == "2 ?pairend":
                countST += 1
                
            elif str(cell_additional_value) == "3 1pairend":
                countST += 1
            elif str(cell_additional_value) == "3 2pairend":
                countST += 1
            elif str(cell_additional_value) == "3 3pairend":
                countST += 1
            elif str(cell_additional_value) == "3 4pairend":
                countST += 1
            elif str(cell_additional_value) == "3 ?pairend":
                countST += 1
                
            elif str(cell_additional_value) == "4 1pairend":
                countST += 1
            elif str(cell_additional_value) == "4 2pairend":
                countST += 1
            elif str(cell_additional_value) == "4 3pairend":
                countST += 1
            elif str(cell_additional_value) == "4 4pairend":
                countST += 1
            elif str(cell_additional_value) == "4 ?pairend":
                countST += 1
                
            elif str(cell_additional_value) == "5 1pairend":
                countST += 1
            elif str(cell_additional_value) == "5 2pairend":
                countST += 1
            elif str(cell_additional_value) == "5 3pairend":
                countST += 1
            elif str(cell_additional_value) == "5 4pairend":
                countST += 1
            elif str(cell_additional_value) == "5 ?pairend":
                countST += 1
                

    content_to_count = "oshibka   "

    for row in sheet.iter_rows(min_row=7, values_only=True):
        cell_content = row[action_column_index] 
        cell_additional_value = row[animal_column_index]
        cell_time_value = row[time_column_index]
        
        
        if cell_time_value is not None and float(cell_time_value) > end_time:
            break

        if cell_content == content_to_count: 
            if str(cell_additional_value) == "Mstop":
                countMS += 1
            elif str(cell_additional_value) == "Mushla":
                countMS += 1
            elif str(cell_additional_value) == "Mchangedposition":
                countMS += 1


    countKS = countST - countMS
                
    values_to_write = [countMS, countKS, countCS, countCUS]

    start_column = openpyxl.utils.column_index_from_string('M')
    for i, value in enumerate(values_to_write):
        sheet.cell(row=2, column=start_column + i, value=value)

    workbook.save(excel_file)
    print(f"Резултат подсчета сосания записан в верхнюю таблицу сосания.")


    cell_ms = "M2"
    cell_ks = "N2"
    output_cell = "M3"

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    value_ms = float(sheet[cell_ms].value)
    value_ks = float(sheet[cell_ks].value)

    stop = value_ms + value_ks
    
    if stop != 0 and value_ms != 0:
        mstop = float(value_ms * 100 / stop)
        sheet[output_cell] = mstop
        workbook.save(excel_file)
    else:
        sheet[output_cell] = 0
        workbook.save(excel_file)


    output_cell = "N3"

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    value_ms = float(sheet[cell_ms].value)
    value_ks = float(sheet[cell_ks].value)


    stop = value_ms + value_ks

    if stop != 0 and value_ks != 0:
        kstop = float(value_ks * 100 / stop)
        sheet[output_cell] = kstop
        workbook.save(excel_file)
    else:
        sheet[output_cell] = 0
        workbook.save(excel_file)



    print(f"Процентное соотношение окончания сосания записано в верхнюю таблицу сосания.")




    cell_ss = "O2"
    cell_uss = "P2"
    output_cell = "O3"

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    value_ss = float(sheet[cell_ss].value)
    value_uss = float(sheet[cell_uss].value)

    attempt = value_ss + value_uss
    
    if attempt != 0 and value_ss != 0:
        success = float(value_ss * 100 / attempt)
        sheet[output_cell] = success
        workbook.save(excel_file)
    else:
        sheet[output_cell] = 0
        workbook.save(excel_file)


    output_cell = "P3"

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    value_ss = float(sheet[cell_ss].value)
    value_uss = float(sheet[cell_uss].value)

    if attempt != 0 and value_uss != 0:
        unsuccess = float(value_uss * 100 / attempt)
        sheet[output_cell] = unsuccess
        workbook.save(excel_file)
    else:
        sheet[output_cell] = 0
        workbook.save(excel_file)
 
    print(f"Процентное соотношение удачных и неудачных попыток записано в верхнюю таблицу сосания.")


    values_to_write = [count1SS, count2SS, count3SS, count4SS, countNSS]
    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AM{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Кол-во удачных попыток сосания по парам записаны в нижнюю таблицу сосания.")


    values_to_write = [count1USS, count2USS, count3USS, count4USS, countNUSS]
    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AN{start_row + i}'] = value
        
    workbook.save(excel_file)#
    print(f"Кол-во неудачных попыток сосания по парам записаны в нижнюю таблицу сосания.")


    p1 = count1SS + count1USS
    p2 = count2SS + count2USS
    p3 = count3SS + count3USS
    p4 = count4SS + count4USS
    pN = countNSS + countNUSS

    values_to_write = [p1, p2, p3, p4, pN]
    start_row = 6
    for i, value in enumerate(values_to_write):
        sheet[f'AO{start_row + i}'] = value
        
    workbook.save(excel_file)#

    print(f"Кол-во total попыток сосания по парам записаны в нижнюю таблицу сосания.")
    
    

    # 1 KITTEN

    output_cell = "O12"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    print(f"Активный лист: {sheet.title}")

    total_time = 0
    kit1_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit1_combination_aliases = {
        "14", "41", "1441", "4114",
        "15", "51", "1551", "5115",
        "13", "31", "1331", "3113",
        "12", "21", "1221", "2112"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit1_event = (animal in kit1_aliases) or (animal in kit1_combination_aliases) or (animal == "1")
            
            if is_kit1_event:
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        kit1_nevidno = True
                else:
                    kit1_nevidno = False
            
            if kit1_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR6"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 1 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    #2 KITTEN

    output_cell = "O13"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    print(f"Активный лист: {sheet.title}")

    total_time = 0
    kit2_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    kit2_combination_aliases = {
        "12", "21", "1221", "2112",
        "23", "32", "2332", "3223",
        "24", "42", "2442", "4224",
        "25", "52", "2552", "5225"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit2_event = (animal in kit2_aliases) or (animal in kit2_combination_aliases) or (animal == "2")
            
            if is_kit2_event:
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        kit2_nevidno = True
                else:
                    kit2_nevidno = False
            
            if kit2_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR7"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 2 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 3 KITTEN

    output_cell = "O14"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    print(f"Активный лист: {sheet.title}")

    total_time = 0
    kit3_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    kit3_combination_aliases = {
        "13", "31", "1331", "3113",
        "23", "32", "2332", "3223",
        "34", "43", "3443", "4334",
        "35", "53", "3553", "5335"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit3_event = (animal in kit3_aliases) or (animal in kit3_combination_aliases) or (animal == "3")
            
            if is_kit3_event:
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        kit3_nevidno = True
                else:
                    kit3_nevidno = False
            
            if kit3_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR8"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 3 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 4 KITTEN

    output_cell = "O15"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    print(f"Активный лист: {sheet.title}")

    total_time = 0
    kit4_nevidno = True
    simultaneous_nevidno_start_time = 0.0  
    last_valid_time = None

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    kit4_combination_aliases = {
        "14", "41", "1441", "4114",
        "24", "42", "2442", "4224",
        "34", "43", "3443", "4334",
        "45", "54", "4554", "5445"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit4_event = (animal in kit4_aliases) or (animal in kit4_combination_aliases) or (animal == "4")
            
            if is_kit4_event:
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        kit4_nevidno = True
                else:
                    kit4_nevidno = False
            
            if kit4_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR9"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 4 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 5 KITTEN

    output_cell = "O16"
    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active
        
    print(f"Активный лист: {sheet.title}")

    total_time = 0
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    kit5_combination_aliases = {
        "15", "51", "1551", "5115",
        "25", "52", "2552", "5225",
        "35", "53", "3553", "5335",
        "45", "54", "4554", "5445"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit5_event = (animal in kit5_aliases) or (animal in kit5_combination_aliases) or (animal == "5")
            
            if is_kit5_event:
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        kit5_nevidno = True
                else:
                    kit5_nevidno = False
            
            if kit5_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR10"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для котенка 5 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # MOTHER

    output_cell = "O11"
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    total_time = 0
    kitM_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kitM_aliases = {
        "m", "mm", "1m", "m1", "1mm1", "m11m", 
        "mneighbour", "neighbourm", "mneighbourneighbourm", "neighbourmmneighbour",
        "mneighbor", "neighborm", "mneighborneighborm", "neighbormmneighbor",
        "2m", "m2", "2mm2", "m22m",
        "3m", "m3", "3mm3", "m33m",
        "4m", "m4", "4mm4", "m44m",
        "5m", "m5", "5mm5", "m55m",
        "kittenm", "Kittenm", "mkitten", "mKitten",
        "1tail", "1paw", "1ears",
        "2tail", "2paw", "2ears",
        "3tail", "3paw", "3ears",
        "4tail", "4paw", "4ears",
        "5tail", "5paw", "5ears"
    }

    special_actions = {"cocanie"}
    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kitM_event = animal in kitM_aliases or animal == "m"
            
            if is_kitM_event:
                if action in non_interrupting_actions:
                    if action == "nevidno":
                        kitM_nevidno = True
                elif action in special_actions:
                    kitM_nevidno = False
                else:
                    kitM_nevidno = False
            else:
                if action in special_actions:
                    kitM_nevidno = False
            
            if kitM_nevidno:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR11"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время невидно для мамы (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 1 END 2

    output_cell = "O27"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit1_nevidno = True
    kit2_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    both_kits_aliases = {
        "12", "21", "1221", "2112"
    }

    kit1_with_others_aliases = {
        "13", "31", "1331", "3113",
        "14", "41", "1441", "4114",
        "15", "51", "1551", "5115"
    }

    kit2_with_others_aliases = {
        "23", "32", "2332", "3223",
        "24", "42", "2442", "4224",
        "25", "52", "2552", "5225"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit1 = animal in kit1_aliases or animal == "1"
            is_kit2 = animal in kit2_aliases or animal == "2"
            is_both = animal in both_kits_aliases
            is_kit1_with_other = animal in kit1_with_others_aliases
            is_kit2_with_other = animal in kit2_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit1:
                        kit1_nevidno = True
                    if is_kit2:
                        kit2_nevidno = True
                    if is_both:
                        kit1_nevidno = True
                        kit2_nevidno = True
                    if is_kit1_with_other:
                        kit1_nevidno = True
                    if is_kit2_with_other:
                        kit2_nevidno = True
            else:
                if is_kit1:
                    kit1_nevidno = False
                if is_kit2:
                    kit2_nevidno = False
                if is_both:
                    kit1_nevidno = False
                    kit2_nevidno = False
                if is_kit1_with_other:
                    kit1_nevidno = False
                if is_kit2_with_other:
                    kit2_nevidno = False
            
            both_invisible = kit1_nevidno and kit2_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR12"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 1 и 2 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")




    # 1 END 3

    output_cell = "O28"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit1_nevidno = True
    kit3_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    both_kits_aliases = {
        "13", "31", "1331", "3113"
    }

    kit1_with_others_aliases = {
        "12", "21", "1221", "2112",
        "14", "41", "1441", "4114",
        "15", "51", "1551", "5115"
    }

    kit3_with_others_aliases = {
        "23", "32", "2332", "3223",
        "34", "43", "3443", "4334",
        "35", "53", "3553", "5335"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit1 = animal in kit1_aliases or animal == "1"
            is_kit3 = animal in kit3_aliases or animal == "3"
            is_both = animal in both_kits_aliases
            is_kit1_with_other = animal in kit1_with_others_aliases
            is_kit3_with_other = animal in kit3_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit1:
                        kit1_nevidno = True
                    if is_kit3:
                        kit3_nevidno = True
                    if is_both:
                        kit1_nevidno = True
                        kit3_nevidno = True
                    if is_kit1_with_other:
                        kit1_nevidno = True
                    if is_kit3_with_other:
                        kit3_nevidno = True
            else:
                if is_kit1:
                    kit1_nevidno = False
                if is_kit3:
                    kit3_nevidno = False
                if is_both:
                    kit1_nevidno = False
                    kit3_nevidno = False
                if is_kit1_with_other:
                    kit1_nevidno = False
                if is_kit3_with_other:
                    kit3_nevidno = False
            
            both_invisible = kit1_nevidno and kit3_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR13"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 1 и 3 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 1 И 4

    output_cell = "O29"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit1_nevidno = True
    kit4_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    both_kits_aliases = {
        "14", "41", "1441", "4114"
    }

    kit1_with_others_aliases = {
        "12", "21", "1221", "2112",
        "13", "31", "1331", "3113",
        "15", "51", "1551", "5115"
    }

    kit4_with_others_aliases = {
        "24", "42", "2442", "4224",
        "34", "43", "3443", "4334",
        "45", "54", "4554", "5445"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit1 = animal in kit1_aliases or animal == "1"
            is_kit4 = animal in kit4_aliases or animal == "4"
            is_both = animal in both_kits_aliases
            is_kit1_with_other = animal in kit1_with_others_aliases
            is_kit4_with_other = animal in kit4_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit1:
                        kit1_nevidno = True
                    if is_kit4:
                        kit4_nevidno = True
                    if is_both:
                        kit1_nevidno = True
                        kit4_nevidno = True
                    if is_kit1_with_other:
                        kit1_nevidno = True
                    if is_kit4_with_other:
                        kit4_nevidno = True
            else:
                if is_kit1:
                    kit1_nevidno = False
                if is_kit4:
                    kit4_nevidno = False
                if is_both:
                    kit1_nevidno = False
                    kit4_nevidno = False
                if is_kit1_with_other:
                    kit1_nevidno = False
                if is_kit4_with_other:
                    kit4_nevidno = False
            
            both_invisible = kit1_nevidno and kit4_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR14"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 1 и 4 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 1 END 5

    output_cell = "O30"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit1_nevidno = True
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit1_aliases = {
        "1", "11", "1m", "m1", "1mm1", "m11m", "1tail", "1paw", "1ears",
        "1neighbour", "neighbour1", "1neighbourneighbour1", "neighbour11neighbour",
        "1neighbor", "neighbor1", "1neighborneighbor1", "neighbor11neighbor",
        "kitten1", "1kitten",
        "1 1pairtry", "1 2pairtry", "1 3pairtry", "1 4pairtry", "1 ?pairtry",
        "1 1pairtrysuccess", "1 2pairtrysuccess", "1 3pairtrysuccess", "1 4pairtrysuccess", "1 ?pairtrysuccess",
        "1 1pairtryunsuccess", "1 2pairtryunsuccess", "1 3pairtryunsuccess", "1 4pairtryunsuccess", "1 ?pairtryunsuccess",
        "1 1pair", "1 2pair", "1 3pair", "1 4pair", "1 ?pair",
        "1 1pairend", "1 2pairend", "1 3pairend", "1 4pairend", "1 ?pairend"
    }

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    both_kits_aliases = {
        "15", "51", "1551", "5115"
    }

    kit1_with_others_aliases = {
        "12", "21", "1221", "2112",
        "13", "31", "1331", "3113",
        "14", "41", "1441", "4114"
    }

    kit5_with_others_aliases = {
        "25", "52", "2552", "5225",
        "35", "53", "3553", "5335",
        "45", "54", "4554", "5445"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit1 = animal in kit1_aliases or animal == "1"
            is_kit5 = animal in kit5_aliases or animal == "5"
            is_both = animal in both_kits_aliases
            is_kit1_with_other = animal in kit1_with_others_aliases
            is_kit5_with_other = animal in kit5_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit1:
                        kit1_nevidno = True
                    if is_kit5:
                        kit5_nevidno = True
                    if is_both:
                        kit1_nevidno = True
                        kit5_nevidno = True
                    if is_kit1_with_other:
                        kit1_nevidno = True
                    if is_kit5_with_other:
                        kit5_nevidno = True
            else:
                if is_kit1:
                    kit1_nevidno = False
                if is_kit5:
                    kit5_nevidno = False
                if is_both:
                    kit1_nevidno = False
                    kit5_nevidno = False
                if is_kit1_with_other:
                    kit1_nevidno = False
                if is_kit5_with_other:
                    kit5_nevidno = False
            
            both_invisible = kit1_nevidno and kit5_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR15"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 1 и 5 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 2 И 3

    output_cell = "O32"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit2_nevidno = True
    kit3_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    both_kits_aliases = {
        "23", "32", "2332", "3223"
    }

    kit2_with_others_aliases = {
        "12", "21", "1221", "2112",
        "24", "42", "2442", "4224",
        "25", "52", "2552", "5225"
    }

    kit3_with_others_aliases = {
        "13", "31", "1331", "3113",
        "34", "43", "3443", "4334",
        "35", "53", "3553", "5335"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit2 = animal in kit2_aliases or animal == "2"
            is_kit3 = animal in kit3_aliases or animal == "3"
            is_both = animal in both_kits_aliases
            is_kit2_with_other = animal in kit2_with_others_aliases
            is_kit3_with_other = animal in kit3_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit2:
                        kit2_nevidno = True
                    if is_kit3:
                        kit3_nevidno = True
                    if is_both:
                        kit2_nevidno = True
                        kit3_nevidno = True
                    if is_kit2_with_other:
                        kit2_nevidno = True
                    if is_kit3_with_other:
                        kit3_nevidno = True
            else:
                if is_kit2:
                    kit2_nevidno = False
                if is_kit3:
                    kit3_nevidno = False
                if is_both:
                    kit2_nevidno = False
                    kit3_nevidno = False
                if is_kit2_with_other:
                    kit2_nevidno = False
                if is_kit3_with_other:
                    kit3_nevidno = False
            
            both_invisible = kit2_nevidno and kit3_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR16"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 2 и 3 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 2 END 4 

    output_cell = "O33"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit2_nevidno = True
    kit4_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    both_kits_aliases = {
        "24", "42", "2442", "4224"
    }

    kit2_with_others_aliases = {
        "12", "21", "1221", "2112",
        "23", "32", "2332", "3223",
        "25", "52", "2552", "5225"
    }

    kit4_with_others_aliases = {
        "14", "41", "1441", "4114",
        "34", "43", "3443", "4334",
        "45", "54", "4554", "5445"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit2 = animal in kit2_aliases or animal == "2"
            is_kit4 = animal in kit4_aliases or animal == "4"
            is_both = animal in both_kits_aliases
            is_kit2_with_other = animal in kit2_with_others_aliases
            is_kit4_with_other = animal in kit4_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit2:
                        kit2_nevidno = True
                    if is_kit4:
                        kit4_nevidno = True
                    if is_both:
                        kit2_nevidno = True
                        kit4_nevidno = True
                    if is_kit2_with_other:
                        kit2_nevidno = True
                    if is_kit4_with_other:
                        kit4_nevidno = True
            else:
                if is_kit2:
                    kit2_nevidno = False
                if is_kit4:
                    kit4_nevidno = False
                if is_both:
                    kit2_nevidno = False
                    kit4_nevidno = False
                if is_kit2_with_other:
                    kit2_nevidno = False
                if is_kit4_with_other:
                    kit4_nevidno = False
            
            both_invisible = kit2_nevidno and kit4_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR17"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 2 и 4 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 2 END 5

    output_cell = "O34"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit2_nevidno = True
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit2_aliases = {
        "2", "22", "2m", "m2", "2mm2", "m22m", "2tail", "2paw", "2ears",
        "2neighbour", "neighbour2", "2neighbourneighbour2", "neighbour22neighbour",
        "2neighbor", "neighbor2", "2neighborneighbor2", "neighbor22neighbor",
        "kitten2", "2kitten",
        "2 1pairtry", "2 2pairtry", "2 3pairtry", "2 4pairtry", "2 ?pairtry",
        "2 1pairtrysuccess", "2 2pairtrysuccess", "2 3pairtrysuccess", "2 4pairtrysuccess", "2 ?pairtrysuccess",
        "2 1pairtryunsuccess", "2 2pairtryunsuccess", "2 3pairtryunsuccess", "2 4pairtryunsuccess", "2 ?pairtryunsuccess",
        "2 1pair", "2 2pair", "2 3pair", "2 4pair", "2 ?pair",
        "2 1pairend", "2 2pairend", "2 3pairend", "2 4pairend", "2 ?pairend"
    }

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    both_kits_aliases = {
        "25", "52", "2552", "5225"
    }

    kit2_with_others_aliases = {
        "12", "21", "1221", "2112",
        "23", "32", "2332", "3223",
        "24", "42", "2442", "4224"
    }

    kit5_with_others_aliases = {
        "15", "51", "1551", "5115",
        "35", "53", "3553", "5335",
        "45", "54", "4554", "5445"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit2 = animal in kit2_aliases or animal == "2"
            is_kit5 = animal in kit5_aliases or animal == "5"
            is_both = animal in both_kits_aliases
            is_kit2_with_other = animal in kit2_with_others_aliases
            is_kit5_with_other = animal in kit5_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit2:
                        kit2_nevidno = True
                    if is_kit5:
                        kit5_nevidno = True
                    if is_both:
                        kit2_nevidno = True
                        kit5_nevidno = True
                    if is_kit2_with_other:
                        kit2_nevidno = True
                    if is_kit5_with_other:
                        kit5_nevidno = True
            else:
                if is_kit2:
                    kit2_nevidno = False
                if is_kit5:
                    kit5_nevidno = False
                if is_both:
                    kit2_nevidno = False
                    kit5_nevidno = False
                if is_kit2_with_other:
                    kit2_nevidno = False
                if is_kit5_with_other:
                    kit5_nevidno = False
            
            both_invisible = kit2_nevidno and kit5_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR18"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 2 и 5 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 3 END 4 

    output_cell = "O37"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit3_nevidno = True
    kit4_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    both_kits_aliases = {
        "34", "43", "3443", "4334"
    }

    kit3_with_others_aliases = {
        "13", "31", "1331", "3113",
        "23", "32", "2332", "3223",
        "35", "53", "3553", "5335"
    }

    kit4_with_others_aliases = {
        "14", "41", "1441", "4114",
        "24", "42", "2442", "4224",
        "45", "54", "4554", "5445"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit3 = animal in kit3_aliases or animal == "3"
            is_kit4 = animal in kit4_aliases or animal == "4"
            is_both = animal in both_kits_aliases
            is_kit3_with_other = animal in kit3_with_others_aliases
            is_kit4_with_other = animal in kit4_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit3:
                        kit3_nevidno = True
                    if is_kit4:
                        kit4_nevidno = True
                    if is_both:
                        kit3_nevidno = True
                        kit4_nevidno = True
                    if is_kit3_with_other:
                        kit3_nevidno = True
                    if is_kit4_with_other:
                        kit4_nevidno = True
            else:
                if is_kit3:
                    kit3_nevidno = False
                if is_kit4:
                    kit4_nevidno = False
                if is_both:
                    kit3_nevidno = False
                    kit4_nevidno = False
                if is_kit3_with_other:
                    kit3_nevidno = False
                if is_kit4_with_other:
                    kit4_nevidno = False
            
            both_invisible = kit3_nevidno and kit4_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR19"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 3 и 4 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 3 END 5

    output_cell = "O38"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit3_nevidno = True
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit3_aliases = {
        "3", "33", "3m", "m3", "3mm3", "m33m", "3tail", "3paw", "3ears",
        "3neighbour", "neighbour3", "3neighbourneighbour3", "neighbour33neighbour",
        "3neighbor", "neighbor3", "3neighborneighbor3", "neighbor33neighbor",
        "kitten3", "3kitten",
        "3 1pairtry", "3 2pairtry", "3 3pairtry", "3 4pairtry", "3 ?pairtry",
        "3 1pairtrysuccess", "3 2pairtrysuccess", "3 3pairtrysuccess", "3 4pairtrysuccess", "3 ?pairtrysuccess",
        "3 1pairtryunsuccess", "3 2pairtryunsuccess", "3 3pairtryunsuccess", "3 4pairtryunsuccess", "3 ?pairtryunsuccess",
        "3 1pair", "3 2pair", "3 3pair", "3 4pair", "3 ?pair",
        "3 1pairend", "3 2pairend", "3 3pairend", "3 4pairend", "3 ?pairend"
    }

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    both_kits_aliases = {
        "35", "53", "3553", "5335"
    }

    kit3_with_others_aliases = {
        "13", "31", "1331", "3113",
        "23", "32", "2332", "3223",
        "34", "43", "3443", "4334"
    }

    kit5_with_others_aliases = {
        "15", "51", "1551", "5115",
        "25", "52", "2552", "5225",
        "45", "54", "4554", "5445"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit3 = animal in kit3_aliases or animal == "3"
            is_kit5 = animal in kit5_aliases or animal == "5"
            is_both = animal in both_kits_aliases
            is_kit3_with_other = animal in kit3_with_others_aliases
            is_kit5_with_other = animal in kit5_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit3:
                        kit3_nevidno = True
                    if is_kit5:
                        kit5_nevidno = True
                    if is_both:
                        kit3_nevidno = True
                        kit5_nevidno = True
                    if is_kit3_with_other:
                        kit3_nevidno = True
                    if is_kit5_with_other:
                        kit5_nevidno = True
            else:
                if is_kit3:
                    kit3_nevidno = False
                if is_kit5:
                    kit5_nevidno = False
                if is_both:
                    kit3_nevidno = False
                    kit5_nevidno = False
                if is_kit3_with_other:
                    kit3_nevidno = False
                if is_kit5_with_other:
                    kit5_nevidno = False
            
            both_invisible = kit3_nevidno and kit5_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR20"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 3 и 5 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



    # 4 END 5

    output_cell = "O42"

    action_column = 'D'
    animal_column = 'K'
    time_column = 'F'

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    total_time = 0
    kit4_nevidno = True
    kit5_nevidno = True
    simultaneous_nevidno_start_time = 0.0
    last_valid_time = None

    kit4_aliases = {
        "4", "44", "4m", "m4", "4mm4", "m44m", "4tail", "4paw", "4ears",
        "4neighbour", "neighbour4", "4neighbourneighbour4", "neighbour44neighbour",
        "4neighbor", "neighbor4", "4neighborneighbor4", "neighbor44neighbor",
        "kitten4", "4kitten",
        "4 1pairtry", "4 2pairtry", "4 3pairtry", "4 4pairtry", "4 ?pairtry",
        "4 1pairtrysuccess", "4 2pairtrysuccess", "4 3pairtrysuccess", "4 4pairtrysuccess", "4 ?pairtrysuccess",
        "4 1pairtryunsuccess", "4 2pairtryunsuccess", "4 3pairtryunsuccess", "4 4pairtryunsuccess", "4 ?pairtryunsuccess",
        "4 1pair", "4 2pair", "4 3pair", "4 4pair", "4 ?pair",
        "4 1pairend", "4 2pairend", "4 3pairend", "4 4pairend", "4 ?pairend"
    }

    kit5_aliases = {
        "5", "55", "5m", "m5", "5mm5", "m55m", "5tail", "5paw", "5ears",
        "5neighbour", "neighbour5", "5neighbourneighbour5", "neighbour55neighbour",
        "5neighbor", "neighbor5", "5neighborneighbor5", "neighbor55neighbor",
        "kitten5", "5kitten",
        "5 1pairtry", "5 2pairtry", "5 3pairtry", "5 4pairtry", "5 ?pairtry",
        "5 1pairtrysuccess", "5 2pairtrysuccess", "5 3pairtrysuccess", "5 4pairtrysuccess", "5 ?pairtrysuccess",
        "5 1pairtryunsuccess", "5 2pairtryunsuccess", "5 3pairtryunsuccess", "5 4pairtryunsuccess", "5 ?pairtryunsuccess",
        "5 1pair", "5 2pair", "5 3pair", "5 4pair", "5 ?pair",
        "5 1pairend", "5 2pairend", "5 3pairend", "5 4pairend", "5 ?pairend"
    }

    both_kits_aliases = {
        "45", "54", "4554", "5445"
    }

    kit4_with_others_aliases = {
        "14", "41", "1441", "4114",
        "24", "42", "2442", "4224",
        "34", "43", "3443", "4334"
    }

    kit5_with_others_aliases = {
        "15", "51", "1551", "5115",
        "25", "52", "2552", "5225",
        "35", "53", "3553", "5335"
    }

    non_interrupting_actions = {"nevidno", "vokal"}

    for row_index, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            action = str(row[openpyxl.utils.column_index_from_string(action_column) - 1] or "").strip().lower()
            animal = str(row[openpyxl.utils.column_index_from_string(animal_column) - 1] or "").strip().lower()
            time = row[openpyxl.utils.column_index_from_string(time_column) - 1]

            if not action or not animal or time is None:
                continue

            current_time = float(time)
            last_valid_time = current_time
            
            if current_time > end_time:
                break

            is_kit4 = animal in kit4_aliases or animal == "4"
            is_kit5 = animal in kit5_aliases or animal == "5"
            is_both = animal in both_kits_aliases
            is_kit4_with_other = animal in kit4_with_others_aliases
            is_kit5_with_other = animal in kit5_with_others_aliases
            
            if action in non_interrupting_actions:
                if action == "nevidno":
                    if is_kit4:
                        kit4_nevidno = True
                    if is_kit5:
                        kit5_nevidno = True
                    if is_both:
                        kit4_nevidno = True
                        kit5_nevidno = True
                    if is_kit4_with_other:
                        kit4_nevidno = True
                    if is_kit5_with_other:
                        kit5_nevidno = True
            else:
                if is_kit4:
                    kit4_nevidno = False
                if is_kit5:
                    kit5_nevidno = False
                if is_both:
                    kit4_nevidno = False
                    kit5_nevidno = False
                if is_kit4_with_other:
                    kit4_nevidno = False
                if is_kit5_with_other:
                    kit5_nevidno = False
            
            both_invisible = kit4_nevidno and kit5_nevidno
            
            if both_invisible:
                if simultaneous_nevidno_start_time is None:
                    simultaneous_nevidno_start_time = current_time
            else:
                if simultaneous_nevidno_start_time is not None:
                    period_end = min(end_time, current_time)
                    duration = period_end - simultaneous_nevidno_start_time
                    if duration > 0:
                        total_time += duration
                    simultaneous_nevidno_start_time = None
                    
        except (ValueError, TypeError):
            continue

    if simultaneous_nevidno_start_time is not None:
        if last_valid_time is not None:
            period_end = min(end_time, last_valid_time)
            duration = period_end - simultaneous_nevidno_start_time
            if duration > 0:
                total_time += duration
        else:
            total_time += end_time

    sheet[output_cell] = total_time
    workbook.save(excel_file)

    output_cell2 = "AR21"
    sheet[output_cell2] = total_time
    workbook.save(excel_file)
    print(f"Общее время одновременного невидно для котят 4 и 5 (до {end_time} сек): {total_time:.2f} секунд. Результат записан в ячейки {output_cell} и {output_cell2}.")



workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

value_r2 = sheet['R2'].value
for row in range(6, 22):
    cell_ar = sheet[f'AR{row}']
    cell_as = sheet[f'AS{row}']
    
    if cell_ar.value is not None:
        cell_as.value = float(value_r2) - float(cell_ar.value)
    else:
        cell_as.value = None

workbook.save(excel_file)
print(f"Время Observed для каждого котенка, мамы и пар котят посчитано.")


print("ПОДСЧЕТ ОКОНЧЕН")
