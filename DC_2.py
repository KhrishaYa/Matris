# -*- coding: utf-8 -*-
"""
Created on Fri Aug 15 19:18:39 2025

@author: KHRISYA YA
"""



import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side

FOLDER_PATH = r"C:\Users\Lenovo\Desktop\Кошки 2024\пиздец" #УКАЖИТЕ ПУТЬ ДО ВАШЕЙ ПАПКИ (НЕ ДО ФАЙЛА!)
RESULT_FILE = "итоговые_расчеты.xlsx" # ЗАМЕНИТЕ НА НАЗВАНИЕ ИТОГОВАОГО ФАЙЛА
EXCEL_FILE = "итоговые_расчеты.xlsx" # ЗАМЕНИТЕ НА НАЗВАНИЕ ИТОГОВАОГО ФАЙЛА
FILE_NAME = "итоговые_расчеты.xlsx"  # ЗАМЕНИТЕ НА НАЗВАНИЕ ИТОГОВАОГО ФАЙЛА
SHEET_NAME = "Лист1"                 # НИЧЕГО НЕ МЕНЯТЬ
SPECIES = "Lynx" # ПОМЕНЯТЬ НА ВИД ОБСЧИТЫВАЕМОГО ЖИВОТНОГО


TARGET_CELLS_PART1 = ['L2']
TARGET_CELLS_PART2 = ['R8', 'X8', 'Y8', 'Z8', 'AA8', 'AB8', 'AC8', 'AD8', 'AE8']
TARGET_CELLS_PART3 = ['T7', 'U7', 'R7', 'X7', 'Y7', 'Z7', 'AA7', 'AB7', 'AC7', 'AD7', 'AE7']
TARGET_CELLS_PART4 = ['T6', 'U6', 'R6', 'X6', 'Y6', 'Z6', 'AA6', 'AB6', 'AC6', 'AD6', 'AE6']
TARGET_CELLS_PART5 = ['U6', 'U7']
TARGET_CELLS_PART6 = ['T7', 'U7', 'R7', 'X7', 'Y7', 'Z7', 'AA7', 'AB7', 'AC7', 'AD7', 'AE7', 'T6', 'U6', 'R6', 'X6', 'Y6', 'Z6', 'AA6', 'AB6', 'AC6', 'AD6', 'AE6', 'U9', 'U10']
TARGET_CELLS_PART7 = ['T7', 'U10', 'R7', 'X7', 'Y7', 'Z7', 'AA7', 'AB7', 'AC7', 'AD7', 'AE7', 'R8', 'X8', 'Y8', 'Z8', 'AA8', 'AB8', 'AC8', 'AD8', 'AE8', 'R62', 'X62', 'Y62', 'Z62', 'AA62', 'AB62', 'AC62', 'AD62', 'AE62']
TARGET_CELLS_PART8 = ['T6', 'U6', 'R6', 'X6', 'Y6', 'Z6', 'AA6', 'AB6', 'AC6', 'AD6', 'AE6', 'R8', 'X8', 'Y8', 'Z8', 'AA8', 'AB8', 'AC8', 'AD8', 'AE8', 'R68', 'X68', 'Y68', 'Z68', 'AA68', 'AB68', 'AC68', 'AD68', 'AE68']
TARGET_CELLS_PART9 = ['AJ7']
TARGET_CELLS_PART10 = ['AJ9']
TARGET_CELLS_PART11 = ['AJ8']
TARGET_CELLS_PART12 = ['AJ6']
TARGET_CELLS_PART13 = ['V9']
TARGET_CELLS_PART14 = ['V11']
TARGET_CELLS_PART15 = ['O2']
TARGET_CELLS_PART16 = ['P2']
TARGET_CELLS_PART17 = ['L2']
TARGET_CELLS_PART18 = ['R8', 'X8', 'Y8', 'Z8', 'AA8', 'AB8', 'AC8', 'AD8', 'AE8']
TARGET_CELLS_PART19 = ['T7', 'U7', 'R7', 'X7', 'Y7', 'Z7', 'AA7', 'AB7', 'AC7', 'AD7', 'AE7']
TARGET_CELLS_PART20 = ['T6', 'U6', 'R6', 'X6', 'Y6', 'Z6', 'AA6', 'AB6', 'AC6', 'AD6', 'AE6']
TARGET_CELLS_PART21 = ['U6', 'U7', 'U9', 'U10']
TARGET_CELLS_PART22 = ['T7', 'U7', 'R7', 'X7', 'Y7', 'Z7', 'AA7', 'AB7', 'AC7', 'AD7', 'AE7', 'T6', 'U6', 'R6', 'X6', 'Y6', 'Z6', 'AA6', 'AB6', 'AC6', 'AD6', 'AE6', 'U9', 'U10']
TARGET_CELLS_PART23 = ['T7', 'U10', 'R7', 'X7', 'Y7', 'Z7', 'AA7', 'AB7', 'AC7', 'AD7', 'AE7', 'R8', 'X8', 'Y8', 'Z8', 'AA8', 'AB8', 'AC8', 'AD8', 'AE8', 'R62', 'X62', 'Y62', 'Z62', 'AA62', 'AB62', 'AC62', 'AD62', 'AE62']
TARGET_CELLS_PART24 = ['T6', 'U6', 'R6', 'X6', 'Y6', 'Z6', 'AA6', 'AB6', 'AC6', 'AD6', 'AE6', 'R8', 'X8', 'Y8', 'Z8', 'AA8', 'AB8', 'AC8', 'AD8', 'AE8', 'R68', 'X68', 'Y68', 'Z68', 'AA68', 'AB68', 'AC68', 'AD68', 'AE68']
TARGET_CELLS_PART25 = ['AJ7']
TARGET_CELLS_PART26 = ['AJ9']
TARGET_CELLS_PART27 = ['AJ8']
TARGET_CELLS_PART28 = ['AJ6']
TARGET_CELLS_PART29 = ['V9']
TARGET_CELLS_PART30 = ['V11']
TARGET_CELLS_PART31 = ['V10']
TARGET_CELLS_PART32 = ['V10']
TARGET_CELLS_PART33 = ['L2']
TARGET_CELLS_PART34 = ['T2']

TARGET_CELLS_PART35 = ['P8']
TARGET_CELLS_PART36 = ['P7']
TARGET_CELLS_PART37 = ['P6']
TARGET_CELLS_PART38 = ['P10']

TARGET_CELLS_PART39 = ['P8']
TARGET_CELLS_PART40 = ['P7']
TARGET_CELLS_PART41 = ['P6']
TARGET_CELLS_PART42 = ['P10']



def apply_border(cell):
    """Применяет рамку к ячейке, если она не пустая."""
    if cell.value is not None:
        cell.border = bold_border

def create_final_table(result_wb):
    """Создает финальную таблицу с заголовками и данными"""
    print("\n=== СОЗДАНИЕ ФИНАЛЬНОЙ ТАБЛИЦЫ ===")
    
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    thin = Side(border_style="thin", color="000000")
    global bold_border
    bold_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    simple_headers = ["Выводок", "Дата", "Вид", "Кол-во к.", "Период", "Наблюдение", "Observed"]
    for idx, header in enumerate(simple_headers, start=1):
        ws.merge_cells(start_row=1, start_column=idx, end_row=3, end_column=idx)
        cell = ws.cell(row=1, column=idx, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(cell)

    col = len(simple_headers) + 1  

    # Многоуровневые секции
    def write_section(title, subblocks):
        nonlocal col
        start_col = col
        for block in subblocks:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
            cell = ws.cell(row=2, column=col, value=block)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            apply_border(cell)
            
            for i, sub in enumerate(["факт", "норм."], start=0):
                cell = ws.cell(row=3, column=col + i, value=sub)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                apply_border(cell)
            col += 2
        
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col-1)
        header_cell = ws.cell(row=1, column=start_col, value=title)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(header_cell)

    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)
    header_cell = ws.cell(row=1, column=col, value="Сосание")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    apply_border(header_cell)

    for i, sub in enumerate(["факт", "норм.", "успех", "неуспех"]):
        ws.merge_cells(start_row=2, start_column=col+i, end_row=3, end_column=col+i)
        cell = ws.cell(row=2, column=col+i, value=sub)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(cell)
    col += 4

    write_section("Игра сибсов", ["KitKit (Total)"])
    write_section("Игра с мамой", ["KitM", "Mkit", "Игра с пас. мамой", "Total"])
    write_section("Kit total", ["initiator", "recipient"])
    write_section("Борьба", ["KitKit", "KitM", "Mkit", "Total"])
    write_section("Игра с предметом", ["Kit", "M", "Total"])
    write_section("Аллогруминг", ["KitKit", "KitM", "Mkit", "Total"])

    result_ws = result_wb.active
    for row in result_ws.iter_rows(min_row=1, values_only=True):
        ws.append(row)
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            apply_border(cell)

    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 14

    wb.save(FILE_NAME)
    print(f"✅ Таблица сохранена в файл '{FILE_NAME}' на лист '{SHEET_NAME}'")
    print(f" Путь к файлу: {os.path.abspath(FILE_NAME)}")




def main():
    print("=== Обработка данных ===")
    
    try:
        groups = group_files(FOLDER_PATH)
        if not groups:
            print("Файлы не найдены!")
            return

        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active
        result_ws.title = "Результаты"
                
        print("\n Вычисление результатов (часть 1)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  
                result_part1 = process_pair_part1(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part1 is not None:
                    result_ws.cell(row=row_num, column=1, value=group_name)
                    result_ws.cell(row=row_num, column=2, value=date.strftime('%d.%m.%Y'))
                    result_ws.cell(row=row_num, column=9, value=result_part1)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part1}")
                    row_num += 1

        print("\n Вычисление результатов (часть 2)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  
                result_part2 = process_pair_part2(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part2 is not None:
                    result_ws.cell(row=row_num, column=13, value=result_part2)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part2}")
                    row_num += 1

        # ЧАСТЬ 3
        print("\n Вычисление результатов (часть 3)...")
        row_num = 1 
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  
                result_part3 = process_pair_part3(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part3 is not None:
                    result_ws.cell(row=row_num, column=15, value=result_part3)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part3}")
                    row_num += 1

        # ЧАСТЬ 4
        print("\n Вычисление результатов (часть 4)...")
        row_num = 1  
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  
                result_part4 = process_pair_part4(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part4 is not None:
                    result_ws.cell(row=row_num, column=17, value=result_part4)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part4}")
                    row_num += 1
                    
        # ЧАСТЬ 5
        print("\n Вычисление результатов (часть 5)...")
        row_num = 1  
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  
                result_part5 = process_pair_part5(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part5 is not None:
                    result_ws.cell(row=row_num, column=19, value=result_part5)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part5}")
                    row_num += 1

        # ЧАСТЬ 6
        print("\n Вычисление результатов (часть 6)...")
        row_num = 1  
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  
                result_part6 = process_pair_part6(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part6 is not None:
                    result_ws.cell(row=row_num, column=21, value=result_part6)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part6}")
                    row_num += 1

        # ЧАСТЬ 7
        print("\n Вычисление результатов (часть 7)...")
        row_num = 1  
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  
                result_part7 = process_pair_part7(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part7 is not None:
                    result_ws.cell(row=row_num, column=23, value=result_part7)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part7}")
                    row_num += 1

        # ЧАСТЬ 8
        print("\n Вычисление результатов (часть 8)...")
        row_num = 1  
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  
                result_part8 = process_pair_part8(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part8 is not None:
                    result_ws.cell(row=row_num, column=25, value=result_part8)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part8}")
                    row_num += 1

        # ЧАСТЬ 9
        print("\n Вычисление результатов (часть 9)...")
        row_num = 1  
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part9 = process_pair_part9(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part9 is not None:
                    result_ws.cell(row=row_num, column=27, value=result_part9)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part9}")
                    row_num += 1

        # ЧАСТЬ 10
        print("\n Вычисление результатов (часть 10)...")
        row_num = 1
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part10 = process_pair_part10(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part10 is not None:
                    result_ws.cell(row=row_num, column=29, value=result_part10)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part10}")
                    row_num += 1

        # ЧАСТЬ 11
        print("\n Вычисление результатов (часть 11)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  
                result_part11 = process_pair_part11(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part11 is not None:
                    result_ws.cell(row=row_num, column=31, value=result_part11)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part11}")
                    row_num += 1

        # ЧАСТЬ 12
        print("\n Вычисление результатов (часть 12)...")
        row_num = 1  
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part12 = process_pair_part12(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part12 is not None:
                    result_ws.cell(row=row_num, column=33, value=result_part12)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part12}")
                    row_num += 1

        # ЧАСТЬ 13
        print("\n Вычисление результатов (часть 13)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part13 = process_pair_part13(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part13 is not None:
                    result_ws.cell(row=row_num, column=35, value=result_part13)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part13}")
                    row_num += 1

        # ЧАСТЬ 14
        print("\n Вычисление результатов (часть 14)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part14 = process_pair_part14(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part14 is not None:
                    result_ws.cell(row=row_num, column=37, value=result_part14)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part14}")
                    row_num += 1

        # ЧАСТЬ 15
        print("\n Вычисление результатов (часть 15)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part15 = process_pair_part15(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part15 is not None:
                    result_ws.cell(row=row_num, column=10, value=result_part15)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part15}")
                    row_num += 1

        # ЧАСТЬ 16
        print("\n Вычисление результатов (часть 16)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part16 = process_pair_part16(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part16 is not None:
                    result_ws.cell(row=row_num, column=11, value=result_part16)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part16}")
                    row_num += 1

        # ЧАСТЬ 17
        print("\n Вычисление результатов (часть 17)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part17 = process_pair_part17(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part17 is not None:
                    result_ws.cell(row=row_num, column=8, value=result_part17)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part17}")
                    row_num += 1
                    
        # ЧАСТЬ 18
        print("\n Вычисление результатов (часть 18)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part18 = process_pair_part18(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part18 is not None:
                    result_ws.cell(row=row_num, column=12, value=result_part18)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part18}")
                    row_num += 1

        # ЧАСТЬ 19
        print("\n Вычисление результатов (часть 19)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part19 = process_pair_part19(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part19 is not None:
                    result_ws.cell(row=row_num, column=14, value=result_part19)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part19}")
                    row_num += 1

        # ЧАСТЬ 20
        print("\n Вычисление результатов (часть 20)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part20 = process_pair_part20(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part20 is not None:
                    result_ws.cell(row=row_num, column=16, value=result_part20)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part20}")
                    row_num += 1

        # ЧАСТЬ 21
        print("\n Вычисление результатов (часть 21)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part21 = process_pair_part21(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part21 is not None:
                    result_ws.cell(row=row_num, column=18, value=result_part21)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part21}")
                    row_num += 1

        # ЧАСТЬ 22
        print("\n Вычисление результатов (часть 22)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part22 = process_pair_part22(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part22 is not None:
                    result_ws.cell(row=row_num, column=20, value=result_part22)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part22}")
                    row_num += 1

        # ЧАСТЬ 23
        print("\n Вычисление результатов (часть 23)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part23 = process_pair_part23(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part23 is not None:
                    result_ws.cell(row=row_num, column=22, value=result_part23)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part23}")
                    row_num += 1

        # ЧАСТЬ 24
        print("\n Вычисление результатов (часть 24)...")
        row_num = 1
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]
                result_part24 = process_pair_part24(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part24 is not None:
                    result_ws.cell(row=row_num, column=24, value=result_part24)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part24}")
                    row_num += 1

        # ЧАСТЬ 25
        print("\n Вычисление результатов (часть 25)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part25 = process_pair_part25(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part25 is not None:
                    result_ws.cell(row=row_num, column=26, value=result_part25)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part25}")
                    row_num += 1

        # ЧАСТЬ 26
        print("\n Вычисление результатов (часть 26)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part26 = process_pair_part26(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part26 is not None:
                    result_ws.cell(row=row_num, column=28, value=result_part26)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part26}")
                    row_num += 1

        # ЧАСТЬ 27
        print("\n Вычисление результатов (часть 27)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part27 = process_pair_part27(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part27 is not None:
                    result_ws.cell(row=row_num, column=30, value=result_part27)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part27}")
                    row_num += 1

        # ЧАСТЬ 28
        print("\n Вычисление результатов (часть 28)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part28 = process_pair_part28(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part28 is not None:
                    result_ws.cell(row=row_num, column=32, value=result_part28)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part28}")
                    row_num += 1

        # ЧАСТЬ 29
        print("\n Вычисление результатов (часть 29)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part29 = process_pair_part29(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part29 is not None:
                    result_ws.cell(row=row_num, column=34, value=result_part29)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part29}")
                    row_num += 1

        # ЧАСТЬ 30
        print("\n Вычисление результатов (часть 30)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part30 = process_pair_part30(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part30 is not None:
                    result_ws.cell(row=row_num, column=36, value=result_part30)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part30}")
                    row_num += 1

        # ЧАСТЬ 31
        print("\n Вычисление результатов (часть 31)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part31 = process_pair_part31(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part31 is not None:
                    result_ws.cell(row=row_num, column=39, value=result_part31)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part31}")
                    row_num += 1

        # ЧАСТЬ 32
        print("\n Вычисление результатов (часть 32)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part32 = process_pair_part32(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part32 is not None:
                    result_ws.cell(row=row_num, column=38, value=result_part32)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part32}")
                    row_num += 1

        # ЧАСТЬ 33
        print("\n Вычисление результатов (часть 33)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part33 = process_pair_part33(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part33 is not None:
                    result_ws.cell(row=row_num, column=9, value=result_part33)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part33}")
                    row_num += 1

        # ЧАСТЬ 34
        print("\n Вычисление результатов (часть 34)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part34 = process_pair_part34(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part34 is not None:
                    result_ws.cell(row=row_num, column=7, value=result_part34)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part34}")
                    row_num += 1




        # ЧАСТЬ 35
        print("\n Вычисление результатов (часть 35)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part35 = process_pair_part35(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part35 is not None:
                    result_ws.cell(row=row_num, column=41, value=result_part35)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part35}")
                    row_num += 1




        # ЧАСТЬ 36
        print("\n Вычисление результатов (часть 36)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part36 = process_pair_part36(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part36 is not None:
                    result_ws.cell(row=row_num, column=43, value=result_part36)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part36}")
                    row_num += 1




        # ЧАСТЬ 37
        print("\n Вычисление результатов (часть 37)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part37 = process_pair_part37(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part37 is not None:
                    result_ws.cell(row=row_num, column=45, value=result_part37)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part37}")
                    row_num += 1




        # ЧАСТЬ 38
        print("\n Вычисление результатов (часть 38)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part38 = process_pair_part38(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part38 is not None:
                    result_ws.cell(row=row_num, column=47, value=result_part38)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part38}")
                    row_num += 1




        # ЧАСТЬ 39
        print("\n Вычисление результатов (часть 39)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part39 = process_pair_part39(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part39 is not None:
                    result_ws.cell(row=row_num, column=40, value=result_part39)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part39}")
                    row_num += 1




        # ЧАСТЬ 40
        print("\n Вычисление результатов (часть 40)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part40 = process_pair_part40(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part40 is not None:
                    result_ws.cell(row=row_num, column=42, value=result_part40)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part40}")
                    row_num += 1




        # ЧАСТЬ 41
        print("\n Вычисление результатов (часть 41)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part41 = process_pair_part41(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part41 is not None:
                    result_ws.cell(row=row_num, column=44, value=result_part41)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part41}")
                    row_num += 1




        # ЧАСТЬ 42
        print("\n Вычисление результатов (часть 42)...")
        row_num = 1  # Сбрасываем счетчик строк
        
        for group_name, files in groups.items():
            print(f"\nОбработка группы: {group_name}")
            
            sorted_files = sorted(files, key=lambda x: x[0])
            
            for i in range(0, len(sorted_files), 2):
                if i+1 >= len(sorted_files):
                    print(f"⚠ Нет пары для файла: {sorted_files[i][1]}")
                    continue
                
                file1, file2 = sorted_files[i][1], sorted_files[i+1][1]
                date = sorted_files[i][0]  # Дата из первого файла пары
                result_part42 = process_pair_part42(
                    os.path.join(FOLDER_PATH, file1),
                    os.path.join(FOLDER_PATH, file2)
                )
                
                if result_part42 is not None:
                    result_ws.cell(row=row_num, column=46, value=result_part42)
                    print(f"  {date.strftime('%d.%m.%Y')}: {result_part42}")
                    row_num += 1





        result_wb.save(RESULT_FILE)
        print(f"\n Промежуточные результаты сохранены в: {os.path.abspath(RESULT_FILE)}")
        
        create_final_table(result_wb)
        
        result_wb.close()
        
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {str(e)}")

def group_files(folder_path):
    """Группирует файлы по названию и году"""
    groups = {}
    
    for file in os.listdir(folder_path):
        if not file.lower().endswith('.xlsx'):
            continue
            
        try:
            parts = file.split('_')
            prefix = parts[0]
            day, month = parts[-3], parts[-2]
            year = '20' + parts[-1].split('.')[0][-3:-1]
            date = datetime.strptime(f"{day}_{month}_{year}", "%d_%m_%Y")
            
            group_key = f"{prefix} (год {year})"
            if group_key not in groups:
                groups[group_key] = []
            groups[group_key].append((date, file))
            
        except Exception as e:
            print(f"Ошибка обработки файла {file}: {str(e)}")
    
    return groups

def process_pair_part1(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 1"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART1
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part2(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 2"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART2
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part3(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 3"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART3
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part4(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 4"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART4
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part5(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 5"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART5
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part6(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 6"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART6
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part7(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 7"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART7
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part8(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 8"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART8
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part9(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 9"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART9
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part10(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 10"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART10
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part11(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 11"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART11
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part12(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 12"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART12
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part13(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 13"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART13
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part14(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 14"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART14
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part15(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 15"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART15
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part16(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 16"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART16
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part17(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 17"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART17
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part18(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 18"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART18
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part19(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 19"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART19
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part20(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 20"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART20
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part21(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 21"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART21
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part22(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 22"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART22
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part23(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 23"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART23
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part24(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 24"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART24
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part25(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 25"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART25
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part26(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 26"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART26
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part27(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 27"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART27
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part28(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 28"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART28
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part29(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 29"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART29
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part30(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 30"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART30
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part31(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 31"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART31
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
def process_pair_part32(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 32"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART32
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()

def process_pair_part33(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 33"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART33
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()
        
        
        
        
def process_pair_part34(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 34"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART34
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()




def process_pair_part35(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 33"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART35
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()




def process_pair_part36(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 33"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART36
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()




def process_pair_part37(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 33"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART37
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()





def process_pair_part38(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 33"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART38
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (14400 * sum1) / sum2
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()




def process_pair_part39(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 34"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART39
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()




def process_pair_part40(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 34"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART40
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()




def process_pair_part41(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 34"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART41
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()



def process_pair_part42(file1_path, file2_path):
    """Обрабатывает пару файлов и возвращает результат для части 34"""
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        
        sum1 = sum(
            get_cell_value(ws1, cell) + get_cell_value(ws2, cell) 
            for cell in TARGET_CELLS_PART42
        )
        
        sum2 = get_cell_value(ws1, 'T2') + get_cell_value(ws2, 'T2')
        
        if sum2 == 0:
            print(f"Деление на 0 в паре: {file1_path} + {file2_path}")
            return 0.0
            
        result = (1 * sum1) / 1
        return round(result, 2)
        
    except Exception as e:
        print(f"Ошибка обработки пары {file1_path} + {file2_path}: {str(e)}")
        return None
    finally:
        wb1.close()
        wb2.close()




def get_cell_value(ws, cell):
    """Безопасно получает значение ячейки (возвращает 0 при ошибке)"""
    try:
        value = ws[cell].value
        return float(value) if value is not None else 0.0
    except:
        return 0.0

if __name__ == "__main__":
    main()
    
print("ПОДСЧЕТ ОКОНЧЕН")

def main():
    print("=== Обработка данных ===")
    
    try:
        print("\n🔍 Анализ файлов в папке...")
        groups = process_files(FOLDER_PATH)
        if not groups:
            print("Не найдено файлов для обработки!")
            return
        
        print("\n=== СПИСОК ГРУПП ===")
        print_groups_info(groups)
        
        wb, ws = init_excel_file(EXCEL_FILE)
        
        group_values = get_group_values_interactive(groups)
        write_species(ws, SPECIES, len(group_values))
        write_group_values_sequential(ws, group_values)
        write_sequential_pairs(ws, len(group_values), column=5)
        write_alternating_numbers(ws, len(group_values), column=6)
        
        wb.save(EXCEL_FILE)
        print_final_report(groups, group_values)
        
    except Exception as e:
        print(f"\n❌ Ошибка: {str(e)}")

def print_groups_info(groups):
    """Выводит список групп с файлами"""
    for i, (group_key, files) in enumerate(groups, 1):
        print(f"\nГруппа {i}: {group_key[0]} (год {group_key[1]})")
        for j, (date, file) in enumerate(sorted(files, key=lambda x: x[0]), 1):
            print(f"  {j}. {file} ({date.strftime('%d.%m.%Y')})")

def init_excel_file(file_path):
    """Инициализирует Excel-файл"""
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        print(f"\nФайл {file_path} загружен")
    else:
        wb = Workbook()
        wb.active.title = "Данные"
        print(f"\nСоздан новый файл: {file_path}")
    return wb, wb.active

def process_files(folder_path):
    """Анализирует файлы и возвращает сгруппированные данные"""
    groups = {}
    
    for file in os.listdir(folder_path):
        if not file.lower().endswith('.xlsx'):
            continue
            
        try:
            parts = file.split('_')
            if len(parts) < 4:
                continue
                
            prefix, day, month = parts[0], parts[-3], parts[-2]
            year = '20' + parts[-1].split('.')[0][-3:-1]
            
            date = datetime.strptime(f"{day}_{month}_{year}", "%d_%m_%Y")
            group_key = (prefix, year)
            
            if group_key not in groups:
                groups[group_key] = []
            groups[group_key].append((date, file))
            
        except Exception as e:
            print(f"  [!] Ошибка в файле {file}: {e}")
            continue
    
    return sorted(groups.items(), key=lambda x: (x[0][0], x[0][1]))

def write_species(ws, species, group_count):
    """Записывает вид в столбец C"""
    print("\nЗапись вида...")
    count = 0
    for row in range(4, 4 + group_count * 12):
        ws.cell(row=row, column=3, value=species)
        count += 1
    print(f"Записано '{species}' в {count} строк столбца C")

def get_group_values_interactive(groups):
    """Интерактивный ввод значений для групп"""
    print("\nВвод значений для групп:")
    group_values = {}
    for i, (group_key, files) in enumerate(groups, 1):
        group_name = f"{group_key[0]} (год {group_key[1]})"
        while True:
            try:
                value = int(input(f"  Группа {i} [{group_name}]: "))
                group_values[i] = value
                break
            except ValueError:
                print("    Ошибка! Введите целое число.")
    return group_values

def write_group_values_sequential(ws, group_values):
    """Записывает значения групп в столбец D (по 12 значений на группу)"""
    print("\nСохранение значений групп в столбец D...")
    row = 4  
    for group_num in sorted(group_values.keys()):
        value = group_values[group_num]
        for _ in range(12):  
            ws.cell(row=row, column=4, value=value)
            row += 1
    print(f"Записано {len(group_values)*12} значений в столбец D")  

def write_sequential_pairs(ws, group_count, column=5):
    """Записывает в столбец E пары чисел от 1 до 6 для каждой группы"""
    print(f"\nГенерация парных чисел 1-6 в столбец {get_column_letter(column)}...")
    base_sequence = []
    for num in range(1, 7):  
        base_sequence.extend([num, num])  
    
    full_sequence = base_sequence * group_count  
    
    for row, num in enumerate(full_sequence, start=4):
        if row > 4 + group_count * 12:  
            break
        ws.cell(row=row, column=column, value=num)
    
    print(f"Записано {len(full_sequence)} значений (112233445566...) в столбец {get_column_letter(column)}")

def write_alternating_numbers(ws, group_count, column=6):
    """Записывает чередующиеся 1-2 в указанный столбец"""
    print("\nГенерация чередующихся чисел...")
    numbers = [1 if i % 2 == 0 else 2 for i in range(group_count * 12)]
    
    for row, num in enumerate(numbers, start=4):
        if row > 4 + group_count * 12:
            break
        ws.cell(row=row, column=column, value=num)
    print(f"Записано {len(numbers)} значений (1,2,1,2...) в столбец {get_column_letter(column)}")

def print_final_report(groups, group_values):
    """Выводит итоговый отчет"""
    print("\n=== ИТОГОВЫЙ ОТЧЕТ ===")
    print(f"\nОбработано групп: {len(groups)}")
    print("\nВведенные значения:")
    for group_num, value in group_values.items():
        print(f"  Группа {group_num}: {value}")
    
    print(f"\n✅ Все данные сохранены в файл: {os.path.abspath(EXCEL_FILE)}")

def get_column_letter(column_index):
    """Преобразует номер столбца в букву"""
    return chr(64 + column_index)

if __name__ == "__main__":
    main()
